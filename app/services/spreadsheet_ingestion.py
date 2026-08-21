"""Best-effort, deterministic spreadsheet structure and visual ingestion.

Excel cells and native chart references are treated as authoritative.  The LLM
is only used to identify logical tables and add semantic descriptions to the
objects discovered here.
"""

from __future__ import annotations

import base64
import inspect
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Mapping

import openpyxl
import pandas as pd
from openai import OpenAI
from openpyxl.utils.cell import get_column_letter, range_boundaries

from app.settings import settings


logger = logging.getLogger(__name__)

MAX_SPREADSHEET_LLM_CHARS = settings.max_spreadsheet_llm_chars
MAX_VISUALS_FOR_LLM = settings.max_visuals_for_llm
MAX_CHART_REFERENCE_CELLS = settings.max_chart_reference_cells
MAX_TABLES_FROM_LLM = settings.max_tables_from_llm
MAX_HEADERS_PER_TABLE = settings.max_headers_per_table
MAX_FINDINGS_FROM_LLM = settings.max_findings_from_llm
SPREADSHEET_LLM_MODEL = settings.spreadsheet_llm_model
SPREADSHEET_VISION_MODEL = settings.spreadsheet_vision_model
MAX_LLM_OUTPUT_TOKENS = settings.max_llm_output_tokens

_MODERN_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
_SUPPORTED_SERIES_TYPES = {"area", "bar", "line", "pie", "scatter"}
_A1_REFERENCE_RE = re.compile(
    r"^(?:'(?P<quoted>(?:[^']|'')+)'|(?P<plain>[^' !^]+))!"
    r"(?P<cells>\$?[A-Za-z]{1,3}\$?\d+"
    r"(?::\$?[A-Za-z]{1,3}\$?\d+)?)$"
)
_DEFINED_NAME_RE = re.compile(r"^[A-Za-z_\\][A-Za-z0-9_.\\]*$")


class SpreadsheetReferenceError(ValueError):
    """Raised when a chart source reference cannot be resolved safely."""


@dataclass
class _SourceData:
    reference: str | None = None
    values: dict[int, Any] = field(default_factory=dict)
    statuses: dict[int, str] = field(default_factory=dict)
    indices: set[int] = field(default_factory=set)
    default_status: str = "unavailable"

    def value_at(self, index: int) -> Any:
        return self.values.get(index)

    def status_at(self, index: int) -> str:
        return self.statuses.get(index, self.default_status)


def _empty_analysis() -> dict[str, Any]:
    return {
        "spreadsheet_summary": "",
        "tables": [],
        "visual_semantics": [],
        "key_findings": [],
    }


def _configured_openai_client() -> OpenAI:
    return OpenAI(
        base_url=settings.llm_base_url,
        api_key=settings.openai_api_key,
        timeout=settings.llm_timeout_seconds,
        max_retries=0,
    )


def _safe_close(workbook: Any) -> None:
    try:
        workbook.close()
        vba_archive = getattr(workbook, "vba_archive", None)
        archive = getattr(workbook, "_archive", None)
        if vba_archive is not None and vba_archive is not archive:
            vba_archive.close()
    except Exception:  # pragma: no cover - close support varies by reader
        logger.debug("Spreadsheet workbook close failed")


def _safe_worksheet_charts(worksheet: Any) -> list[Any]:
    """Isolate openpyxl's private drawing enumeration API.

    openpyxl does not expose a public worksheet chart collection.  Keeping the
    private access here makes a future version-specific adaptation local.
    """

    try:
        return list(getattr(worksheet, "_charts", ()) or ())
    except Exception:
        logger.warning("Spreadsheet chart enumeration failed")
        return []


def _safe_worksheet_images(worksheet: Any) -> list[Any]:
    """Isolate openpyxl's private embedded-image enumeration API."""

    try:
        return list(getattr(worksheet, "_images", ()) or ())
    except Exception:
        logger.warning("Spreadsheet image enumeration failed")
        return []


def _anchor_location(anchor: Any) -> str:
    if isinstance(anchor, str):
        return anchor

    marker = getattr(anchor, "_from", None)
    if marker is not None:
        try:
            return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
        except Exception:
            logger.debug("Spreadsheet visual has an invalid cell anchor")

    position = getattr(anchor, "pos", None)
    if position is not None:
        x = getattr(position, "x", None)
        y = getattr(position, "y", None)
        if x is not None and y is not None:
            return f"Absolute(x={x}, y={y})"

    return "Unknown"


def _cache_points(cache: Any) -> tuple[dict[int, Any], set[int]]:
    values: dict[int, Any] = {}
    indices: set[int] = set()
    if cache is None:
        return values, indices

    for position, point in enumerate(getattr(cache, "pt", ()) or ()):
        try:
            index = int(getattr(point, "idx", position))
        except (TypeError, ValueError):
            index = position
        values[index] = getattr(point, "v", None)
        indices.add(index)

    point_count = getattr(cache, "ptCount", None)
    if isinstance(point_count, int) and 0 <= point_count <= MAX_CHART_REFERENCE_CELLS:
        indices.update(range(point_count))
    return values, indices


def _multilevel_cache_points(cache: Any) -> tuple[dict[int, Any], set[int]]:
    by_index: dict[int, list[str]] = {}
    indices: set[int] = set()
    if cache is None:
        return {}, set()

    for level in getattr(cache, "lvl", ()) or ():
        level_values, level_indices = _cache_points(level)
        indices.update(level_indices)
        for index, value in level_values.items():
            if value is not None:
                by_index.setdefault(index, []).append(str(value))

    values = {
        index: " / ".join(parts) if parts else None
        for index, parts in by_index.items()
    }
    point_count = getattr(cache, "ptCount", None)
    if isinstance(point_count, int) and 0 <= point_count <= MAX_CHART_REFERENCE_CELLS:
        indices.update(range(point_count))
    return values, indices


def _strict_a1_reference(
    reference: str,
    workbook: Any,
) -> tuple[str, tuple[int, int, int, int]]:
    raw_reference = reference.strip()
    if raw_reference.startswith("="):
        raw_reference = raw_reference[1:].strip()

    match = _A1_REFERENCE_RE.fullmatch(raw_reference)
    if not match:
        raise SpreadsheetReferenceError(
            f"Unsupported or non-A1 chart reference: {reference!r}"
        )

    sheet_name = match.group("quoted") or match.group("plain") or ""
    sheet_name = sheet_name.replace("''", "'")
    worksheet_names = {worksheet.title for worksheet in workbook.worksheets}
    if sheet_name not in worksheet_names:
        raise SpreadsheetReferenceError(
            f"Referenced worksheet does not exist: {sheet_name!r}"
        )

    try:
        min_col, min_row, max_col, max_row = range_boundaries(
            match.group("cells")
        )
    except (TypeError, ValueError) as exc:
        raise SpreadsheetReferenceError(str(exc)) from exc

    if not all((min_col, min_row, max_col, max_row)):
        raise SpreadsheetReferenceError("Whole-row/column references are unsupported")
    if min_col > max_col or min_row > max_row:
        raise SpreadsheetReferenceError("Chart range bounds are reversed")
    if max_col > 16384 or max_row > 1_048_576:
        raise SpreadsheetReferenceError("Chart range exceeds Excel worksheet bounds")

    cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
    if cell_count > MAX_CHART_REFERENCE_CELLS:
        raise SpreadsheetReferenceError(
            f"Chart range contains {cell_count} cells; limit is "
            f"{MAX_CHART_REFERENCE_CELLS}"
        )
    return sheet_name, (min_col, min_row, max_col, max_row)


def _defined_name_target(
    reference: str,
    workbook: Any,
    context_sheet_name: str | None,
) -> tuple[str, tuple[int, int, int, int]]:
    raw_reference = reference.strip().lstrip("=")
    scope_sheet = context_sheet_name
    name = raw_reference

    worksheet_names = {worksheet.title for worksheet in workbook.worksheets}
    if "!" in raw_reference:
        scope, name = raw_reference.rsplit("!", 1)
        if scope.startswith("'") and scope.endswith("'"):
            scope = scope[1:-1].replace("''", "'")
        if scope not in worksheet_names:
            raise SpreadsheetReferenceError(
                f"Named-range scope does not exist: {scope!r}"
            )
        scope_sheet = scope

    if not _DEFINED_NAME_RE.fullmatch(name):
        raise SpreadsheetReferenceError(
            f"Unsupported defined-name reference: {reference!r}"
        )

    definition = None
    if scope_sheet and scope_sheet in worksheet_names:
        definition = workbook[scope_sheet].defined_names.get(name)
    if definition is None:
        definition = workbook.defined_names.get(name)
    if definition is None:
        raise SpreadsheetReferenceError(f"Defined name not found: {name!r}")
    if getattr(definition, "is_external", False):
        raise SpreadsheetReferenceError("External defined names are unsupported")
    if getattr(definition, "type", None) != "RANGE":
        raise SpreadsheetReferenceError("Formula-defined names are unsupported")

    destinations = list(definition.destinations)
    if len(destinations) != 1:
        raise SpreadsheetReferenceError(
            "Defined names must resolve to exactly one cell range"
        )
    target_sheet, cells = destinations[0]
    target_sheet = target_sheet.replace("''", "'")
    quoted = target_sheet.replace("'", "''")
    return _strict_a1_reference(f"'{quoted}'!{cells}", workbook)


def _resolve_reference(
    reference: str,
    formula_workbook: Any,
    value_workbook: Any,
    *,
    context_sheet_name: str | None,
) -> list[tuple[Any, Any]]:
    try:
        sheet_name, bounds = _strict_a1_reference(reference, formula_workbook)
    except SpreadsheetReferenceError as direct_error:
        try:
            sheet_name, bounds = _defined_name_target(
                reference,
                formula_workbook,
                context_sheet_name,
            )
        except SpreadsheetReferenceError:
            raise direct_error

    min_col, min_row, max_col, max_row = bounds
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    cells: list[tuple[Any, Any]] = []
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cells.append(
                (
                    formula_sheet.cell(row=row, column=column),
                    value_sheet.cell(row=row, column=column),
                )
            )
    return cells


def _resolve_reference_grid(
    reference: str,
    formula_workbook: Any,
    value_workbook: Any,
    *,
    context_sheet_name: str | None,
) -> list[list[tuple[Any, Any]]]:
    try:
        sheet_name, bounds = _strict_a1_reference(reference, formula_workbook)
    except SpreadsheetReferenceError as direct_error:
        try:
            sheet_name, bounds = _defined_name_target(
                reference,
                formula_workbook,
                context_sheet_name,
            )
        except SpreadsheetReferenceError:
            raise direct_error

    min_col, min_row, max_col, max_row = bounds
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    return [
        [
            (
                formula_sheet.cell(row=row, column=column),
                value_sheet.cell(row=row, column=column),
            )
            for column in range(min_col, max_col + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]


def _is_formula_cell(cell: Any) -> bool:
    value = getattr(cell, "value", None)
    return getattr(cell, "data_type", None) == "f" or (
        isinstance(value, str) and value.startswith("=")
    )


def _multilevel_cell_data(
    grid: list[list[tuple[Any, Any]]],
    cached_values: Mapping[int, Any],
    *,
    expected_point_count: int | None,
) -> _SourceData:
    if not grid or not grid[0]:
        return _SourceData()

    row_count = len(grid)
    column_count = len(grid[0])
    if expected_point_count and row_count == expected_point_count:
        groups = grid
    elif expected_point_count and column_count == expected_point_count:
        groups = [
            [grid[row][column] for row in range(row_count)]
            for column in range(column_count)
        ]
    elif row_count >= column_count:
        groups = grid
    else:
        groups = [
            [grid[row][column] for row in range(row_count)]
            for column in range(column_count)
        ]

    result = _SourceData()
    for index, group in enumerate(groups):
        result.indices.add(index)
        parts: list[str] = []
        statuses: list[str] = []
        unavailable = False
        for formula_cell, value_cell in group:
            value = value_cell.value
            if _is_formula_cell(formula_cell):
                if value is None:
                    unavailable = True
                    statuses.append("unavailable")
                    continue
                statuses.append("formula_cache")
            else:
                statuses.append("cell" if value is not None else "blank")
            if value is not None:
                parts.append(_safe_scalar(value))

        if unavailable:
            if index in cached_values:
                result.values[index] = cached_values[index]
                result.statuses[index] = "chart_cache"
            else:
                result.values[index] = None
                result.statuses[index] = "unavailable"
        else:
            result.values[index] = " / ".join(parts) if parts else None
            result.statuses[index] = (
                "formula_cache"
                if "formula_cache" in statuses
                else "cell"
                if parts
                else "blank"
            )
    return result


def _source_data(
    source: Any,
    formula_workbook: Any,
    value_workbook: Any,
    *,
    context_sheet_name: str | None,
    context: str,
    expected_point_count: int | None = None,
) -> _SourceData:
    if source is None:
        return _SourceData()

    reference: str | None = None
    cache: Any = None
    literal: Any = None
    is_multilevel = False

    for attribute, cache_attribute in (
        ("multiLvlStrRef", "multiLvlStrCache"),
        ("strRef", "strCache"),
        ("numRef", "numCache"),
    ):
        candidate = getattr(source, attribute, None)
        if candidate is not None:
            reference = getattr(candidate, "f", None)
            cache = getattr(candidate, cache_attribute, None)
            is_multilevel = attribute == "multiLvlStrRef"
            break

    if reference is None:
        for attribute in ("strLit", "numLit"):
            candidate = getattr(source, attribute, None)
            if candidate is not None:
                literal = candidate
                break

    if literal is not None:
        values, indices = _cache_points(literal)
        return _SourceData(
            values=values,
            statuses={index: "literal" for index in indices},
            indices=indices,
            default_status="unavailable",
        )

    if is_multilevel:
        cached_values, cached_indices = _multilevel_cache_points(cache)
    else:
        cached_values, cached_indices = _cache_points(cache)

    result = _SourceData(reference=reference)
    if not reference:
        result.values.update(cached_values)
        result.indices.update(cached_indices)
        result.statuses.update(
            {index: "chart_cache" for index in cached_indices}
        )
        return result

    try:
        grid = _resolve_reference_grid(
            reference,
            formula_workbook,
            value_workbook,
            context_sheet_name=context_sheet_name,
        )
        row_count = len(grid)
        column_count = len(grid[0]) if grid else 0
        rectangular_categories = bool(
            expected_point_count
            and row_count * column_count != expected_point_count
            and expected_point_count in {row_count, column_count}
        )
        if is_multilevel or rectangular_categories:
            multilevel = _multilevel_cell_data(
                grid,
                cached_values,
                expected_point_count=(
                    expected_point_count
                    or (len(cached_indices) if cached_indices else None)
                ),
            )
            multilevel.reference = reference
            return multilevel
        cells = [cell for row in grid for cell in row]
    except SpreadsheetReferenceError:
        logger.warning("Spreadsheet chart reference could not be resolved")
        result.values.update(cached_values)
        result.indices.update(cached_indices)
        result.statuses.update(
            {index: "chart_cache" for index in cached_indices}
        )
        return result

    for index, (formula_cell, value_cell) in enumerate(cells):
        result.indices.add(index)
        value = value_cell.value
        if _is_formula_cell(formula_cell):
            if value is not None:
                result.values[index] = value
                result.statuses[index] = "formula_cache"
            elif index in cached_values:
                result.values[index] = cached_values[index]
                result.statuses[index] = "chart_cache"
            else:
                result.values[index] = None
                result.statuses[index] = "unavailable"
                logger.warning("Spreadsheet formula has no cached chart value")
        else:
            result.values[index] = value
            result.statuses[index] = "cell" if value is not None else "blank"

    return result


def _series_label(
    series: Any,
    series_index: int,
    formula_workbook: Any,
    value_workbook: Any,
    sheet_name: str,
    context: str,
) -> str:
    label = getattr(series, "tx", None)
    reference = getattr(getattr(label, "strRef", None), "f", None)
    if reference:
        cache = getattr(getattr(label, "strRef", None), "strCache", None)
        cached_values, _ = _cache_points(cache)
        try:
            cells = _resolve_reference(
                reference,
                formula_workbook,
                value_workbook,
                context_sheet_name=sheet_name,
            )
            if cells:
                formula_cell, value_cell = cells[0]
                value = value_cell.value
                if value is not None:
                    return str(value)
                if not _is_formula_cell(formula_cell):
                    if formula_cell.value is not None:
                        return str(formula_cell.value)
                    # A resolved blank cell is authoritative over a stale cache.
                    return f"Series {series_index + 1}"
        except SpreadsheetReferenceError:
            logger.warning("Spreadsheet chart series title could not be resolved")
        if cached_values:
            first_index = min(cached_values)
            cached = cached_values[first_index]
            if cached is not None:
                return str(cached)

    literal = getattr(label, "v", None)
    if literal is not None and str(literal).strip():
        return str(literal).strip()
    return f"Series {series_index + 1}"


def _rich_text_value(rich_text: Any) -> str:
    paragraphs: list[str] = []
    for paragraph in getattr(rich_text, "p", ()) or ():
        parts = [
            str(run.t)
            for run in (getattr(paragraph, "r", ()) or ())
            if getattr(run, "t", None) is not None
        ]
        field_value = getattr(getattr(paragraph, "fld", None), "t", None)
        if field_value is not None:
            parts.append(str(field_value))
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _title_value(
    title: Any,
    formula_workbook: Any,
    value_workbook: Any,
    *,
    sheet_name: str,
) -> str:
    if title is None:
        return ""
    if isinstance(title, str):
        return title.strip()

    text = getattr(title, "tx", None)
    string_reference = getattr(text, "strRef", None)
    reference = getattr(string_reference, "f", None)
    if reference:
        try:
            cells = _resolve_reference(
                reference,
                formula_workbook,
                value_workbook,
                context_sheet_name=sheet_name,
            )
            if cells:
                formula_cell, value_cell = cells[0]
                value = value_cell.value
                if value is None and not _is_formula_cell(formula_cell):
                    value = formula_cell.value
                    if value is None:
                        # A resolved blank cell is authoritative over stale title
                        # cache content.
                        return ""
                if value is not None:
                    return str(value).strip()
        except SpreadsheetReferenceError:
            logger.warning("Spreadsheet chart title could not be resolved")
        cached, _ = _cache_points(
            getattr(string_reference, "strCache", None)
        )
        if cached:
            value = cached[min(cached)]
            if value is not None:
                return str(value).strip()

    return _rich_text_value(getattr(text, "rich", None))


def _plot_charts(chart: Any) -> list[Any]:
    try:
        candidates = list(getattr(chart, "_charts", ()) or ()) or [chart]
    except Exception:
        logger.warning("Spreadsheet combined chart enumeration failed")
        candidates = [chart]

    unique: list[Any] = []
    seen: set[int] = set()
    for candidate in candidates:
        identity = id(candidate)
        if identity not in seen:
            seen.add(identity)
            unique.append(candidate)
    return unique


def _chart_series(
    chart: Any,
    formula_workbook: Any,
    value_workbook: Any,
    *,
    sheet_name: str,
    visual_id: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    extracted: list[dict[str, Any]] = []
    chart_types: list[str] = []
    series_index = 0

    for plot in _plot_charts(chart):
        plot_type = type(plot).__name__
        if plot_type not in chart_types:
            chart_types.append(plot_type)
        series_type = str(getattr(plot, "_series_type", "")).lower()
        if series_type not in _SUPPORTED_SERIES_TYPES:
            logger.warning(
                "Unsupported spreadsheet chart plot type=%s series_type=%s",
                plot_type,
                series_type or "unknown",
            )
            continue

        plot_x_axis_title = _title_value(
            getattr(getattr(plot, "x_axis", None), "title", None),
            formula_workbook,
            value_workbook,
            sheet_name=sheet_name,
        )
        plot_y_axis_title = _title_value(
            getattr(getattr(plot, "y_axis", None), "title", None),
            formula_workbook,
            value_workbook,
            sheet_name=sheet_name,
        )

        for series in list(getattr(plot, "ser", ()) or ()):
            context = (
                f"visual_id={visual_id} sheet={sheet_name} "
                f"series={series_index + 1}"
            )
            try:
                name = _series_label(
                    series,
                    series_index,
                    formula_workbook,
                    value_workbook,
                    sheet_name,
                    context,
                )
                item: dict[str, Any] = {
                    "series_index": series_index,
                    "series_name": name,
                    "chart_type": plot_type,
                    "datapoints": [],
                }
                if plot_x_axis_title:
                    item["x_axis_title"] = plot_x_axis_title
                if plot_y_axis_title:
                    item["y_axis_title"] = plot_y_axis_title

                if series_type == "scatter":
                    x_values = _source_data(
                        getattr(series, "xVal", None),
                        formula_workbook,
                        value_workbook,
                        context_sheet_name=sheet_name,
                        context=f"{context} x-values",
                    )
                    y_values = _source_data(
                        getattr(series, "yVal", None),
                        formula_workbook,
                        value_workbook,
                        context_sheet_name=sheet_name,
                        context=f"{context} y-values",
                    )
                    item["x_range"] = x_values.reference
                    item["y_range"] = y_values.reference
                    for point_index in sorted(x_values.indices | y_values.indices):
                        item["datapoints"].append(
                            {
                                "point_index": point_index,
                                "x": x_values.value_at(point_index),
                                "y": y_values.value_at(point_index),
                                "x_status": x_values.status_at(point_index),
                                "y_status": y_values.status_at(point_index),
                            }
                        )
                else:
                    values = _source_data(
                        getattr(series, "val", None),
                        formula_workbook,
                        value_workbook,
                        context_sheet_name=sheet_name,
                        context=f"{context} values",
                    )
                    categories = _source_data(
                        getattr(series, "cat", None),
                        formula_workbook,
                        value_workbook,
                        context_sheet_name=sheet_name,
                        context=f"{context} categories",
                        expected_point_count=(
                            len(values.indices) if values.indices else None
                        ),
                    )
                    item["category_range"] = categories.reference
                    item["value_range"] = values.reference
                    for point_index in sorted(categories.indices | values.indices):
                        item["datapoints"].append(
                            {
                                "point_index": point_index,
                                "category": categories.value_at(point_index),
                                "value": values.value_at(point_index),
                                "category_status": categories.status_at(point_index),
                                "value_status": values.status_at(point_index),
                            }
                        )
                extracted.append(item)
            except Exception:
                logger.warning("Spreadsheet chart series extraction failed")
            finally:
                series_index += 1

    return extracted, chart_types


def _scan_chart(
    chart: Any,
    formula_workbook: Any,
    value_workbook: Any,
    *,
    sheet_name: str,
    visual_id: str,
) -> dict[str, Any]:
    series, chart_types = _chart_series(
        chart,
        formula_workbook,
        value_workbook,
        sheet_name=sheet_name,
        visual_id=visual_id,
    )
    chart_type = type(chart).__name__
    x_axis = getattr(chart, "x_axis", None)
    y_axis = getattr(chart, "y_axis", None)
    result: dict[str, Any] = {
        "visual_id": visual_id,
        "visual_type": "chart",
        "sheet_name": sheet_name,
        "chart_type": chart_type,
        "title": _title_value(
            getattr(chart, "title", None),
            formula_workbook,
            value_workbook,
            sheet_name=sheet_name,
        ),
        "location": _anchor_location(getattr(chart, "anchor", None)),
        "series": series,
    }
    if len(chart_types) > 1:
        result["plot_types"] = chart_types
    if x_axis is not None:
        result["x_axis_title"] = _title_value(
            getattr(x_axis, "title", None),
            formula_workbook,
            value_workbook,
            sheet_name=sheet_name,
        )
    if y_axis is not None:
        result["y_axis_title"] = _title_value(
            getattr(y_axis, "title", None),
            formula_workbook,
            value_workbook,
            sheet_name=sheet_name,
        )
    if chart_type.startswith("Bar"):
        result["bar_direction"] = getattr(chart, "barDir", None)
    return result


def _legacy_sheet_manifest(file_path: str) -> list[dict[str, Any]]:
    sheets: list[dict[str, Any]] = []
    try:
        excel_file = pd.ExcelFile(file_path)
        for sheet_name in excel_file.sheet_names:
            frame = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            sheets.append(
                {
                    "sheet_name": str(sheet_name),
                    "max_row": int(frame.shape[0]),
                    "max_column": int(frame.shape[1]),
                }
            )
    except Exception:
        logger.warning("Legacy spreadsheet scan failed")
    return sheets


def scan_workbook(
    file_path: str,
    *,
    filename: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """Return deterministic sheet, native chart, and image metadata.

    Native drawing extraction is intentionally skipped for legacy ``.xls``;
    pandas remains responsible for its table-compatible cell ingestion.
    """

    display_name = filename or Path(file_path).name
    extension = Path(file_path).suffix.lower()
    if extension == ".xls":
        logger.warning(
            "Legacy XLS chart/image extraction is unsupported; continuing with table ingestion"
        )
        return {
            "sheets": _legacy_sheet_manifest(file_path),
            "charts": [],
            "images": [],
        }
    if extension not in _MODERN_EXCEL_EXTENSIONS:
        raise ValueError(f"Unsupported workbook extension: {extension}")

    keep_vba = extension == ".xlsm"
    formula_workbook = openpyxl.load_workbook(
        file_path,
        data_only=False,
        read_only=False,
        keep_vba=keep_vba,
    )
    try:
        value_workbook = openpyxl.load_workbook(
            file_path,
            data_only=True,
            read_only=False,
            keep_vba=keep_vba,
        )
    except Exception:
        _safe_close(formula_workbook)
        raise

    manifest: dict[str, list[dict[str, Any]]] = {
        "sheets": [],
        "charts": [],
        "images": [],
    }
    try:
        for worksheet in formula_workbook.worksheets:
            sheet_name = worksheet.title
            manifest["sheets"].append(
                {
                    "sheet_name": sheet_name,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                }
            )

            for chart_index, chart in enumerate(_safe_worksheet_charts(worksheet)):
                visual_id = f"{sheet_name}:chart:{chart_index}"
                try:
                    manifest["charts"].append(
                        _scan_chart(
                            chart,
                            formula_workbook,
                            value_workbook,
                            sheet_name=sheet_name,
                            visual_id=visual_id,
                        )
                    )
                except Exception:
                    logger.warning("Spreadsheet chart scan failed")

            for image_index, image in enumerate(_safe_worksheet_images(worksheet)):
                visual_id = f"{sheet_name}:image:{image_index}"
                try:
                    manifest["images"].append(
                        {
                            "visual_id": visual_id,
                            "visual_type": "image",
                            "sheet_name": sheet_name,
                            "location": _anchor_location(
                                getattr(image, "anchor", None)
                            ),
                            "format": getattr(image, "format", None),
                            "width": getattr(image, "width", None),
                            "height": getattr(image, "height", None),
                        }
                    )
                except Exception:
                    logger.warning("Spreadsheet image scan failed")

        # A workbook chartsheet is not part of ``workbook.worksheets``.  It can
        # still own a native chart whose series reference ordinary worksheets.
        for chart_sheet in getattr(formula_workbook, "chartsheets", ()) or ():
            sheet_name = chart_sheet.title
            for chart_index, chart in enumerate(_safe_worksheet_charts(chart_sheet)):
                visual_id = f"{sheet_name}:chart:{chart_index}"
                try:
                    manifest["charts"].append(
                        _scan_chart(
                            chart,
                            formula_workbook,
                            value_workbook,
                            sheet_name=sheet_name,
                            visual_id=visual_id,
                        )
                    )
                except Exception:
                    logger.warning("Spreadsheet chart-sheet scan failed")
    finally:
        _safe_close(value_workbook)
        _safe_close(formula_workbook)
    return manifest


def _safe_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _local_table_range(cell_range: str) -> tuple[int, int, int, int]:
    if not isinstance(cell_range, str) or not cell_range.strip():
        raise ValueError("Table cell_range is required")
    if "!" in cell_range:
        raise ValueError("Table cell_range must not include a worksheet name")
    min_col, min_row, max_col, max_row = range_boundaries(cell_range.strip())
    if not all((min_col, min_row, max_col, max_row)):
        raise ValueError("Whole-row/column table ranges are unsupported")
    if min_col > max_col or min_row > max_row:
        raise ValueError("Table range bounds are reversed")
    return min_col, min_row, max_col, max_row


def chunk_table_deterministically(
    file_path: str,
    table_meta: Mapping[str, Any],
) -> list[str]:
    """Extract one row chunk per table record from the declared worksheet."""

    table_name = str(table_meta.get("table_name") or "Unknown Table").strip()
    sheet_name = str(table_meta.get("sheet_name") or "").strip()
    cell_range = str(table_meta.get("cell_range") or "").strip()
    description = str(table_meta.get("description") or "").strip()
    raw_headers = table_meta.get("column_headers", [])
    headers = (
        [
            str(header).strip()
            if isinstance(header, (str, int, float)) and not pd.isna(header)
            else ""
            for header in raw_headers
        ]
        if isinstance(raw_headers, list)
        else []
    )

    if not sheet_name:
        logger.warning("Spreadsheet table has no sheet name; skipping extraction")
        return []

    try:
        excel_file = pd.ExcelFile(file_path)
        if sheet_name not in excel_file.sheet_names:
            logger.warning("Spreadsheet table references a missing sheet")
            return []
        frame = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    except Exception:
        logger.warning("Spreadsheet table read failed")
        return []

    try:
        min_col, min_row, max_col, max_row = _local_table_range(cell_range)
    except (TypeError, ValueError):
        logger.warning("Spreadsheet table range is invalid")
        return []

    frame_slice = frame.iloc[min_row - 1 : max_row, min_col - 1 : max_col]
    context_prefix = (
        f"[Table: {table_name}]\n"
        f"Description: {description}\n"
        f"Headers: {', '.join(headers)}\n---"
    )
    header_set = {header.casefold() for header in headers if header}
    row_chunks: list[str] = []

    for _, row in frame_slice.iterrows():
        present_values = [value for value in row.values if pd.notna(value)]
        normalized_values = {
            _safe_scalar(value).strip().casefold() for value in present_values
        }
        if header_set and normalized_values:
            matches = len(normalized_values & header_set)
            if matches >= len(header_set) * 0.6:
                continue

        fields: list[str] = []
        for index, value in enumerate(row.values):
            if pd.isna(value):
                continue
            rendered = _safe_scalar(value)
            if index < len(headers) and headers[index]:
                fields.append(f"{headers[index]}: {rendered}")
            else:
                fields.append(rendered)
        if fields:
            row_chunks.append(
                f"{context_prefix}\nRow Data: {' | '.join(fields)}"
            )
    return row_chunks


def _structural_text_modern(file_path: str) -> str:
    extension = Path(file_path).suffix.lower()
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True,
        read_only=True,
        keep_vba=extension == ".xlsm",
    )
    try:
        blocks: list[str] = []
        used = 0
        for worksheet in workbook.worksheets:
            header = (
                f"=== SHEET: {worksheet.title} "
                f"({worksheet.max_row} rows x {worksheet.max_column} columns) ===\n"
            )
            blocks.append(header)
            used += len(header)
            if used >= MAX_SPREADSHEET_LLM_CHARS:
                continue

            for row in worksheet.iter_rows():
                populated = [
                    f"{cell.coordinate}={_safe_scalar(cell.value)}"
                    for cell in row
                    if cell.value is not None
                ]
                if not populated:
                    continue
                line = " | ".join(populated) + "\n"
                remaining = MAX_SPREADSHEET_LLM_CHARS - used
                if remaining <= 0:
                    break
                if len(line) > remaining:
                    blocks.append(line[:remaining])
                    used += remaining
                    break
                blocks.append(line)
                used += len(line)
        if used >= MAX_SPREADSHEET_LLM_CHARS:
            blocks.append("\n[Cell structure truncated at configured limit]\n")
        return "".join(blocks)
    finally:
        _safe_close(workbook)


def _structural_text_legacy(file_path: str) -> str:
    blocks: list[str] = []
    used = 0
    excel_file = pd.ExcelFile(file_path)
    for sheet_name in excel_file.sheet_names:
        frame = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        header = (
            f"=== SHEET: {sheet_name} "
            f"({frame.shape[0]} rows x {frame.shape[1]} columns) ===\n"
        )
        blocks.append(header)
        used += len(header)
        if used >= MAX_SPREADSHEET_LLM_CHARS:
            continue

        for row_number, row in frame.iterrows():
            populated: list[str] = []
            for column_number, value in enumerate(row.values, start=1):
                if pd.isna(value):
                    continue
                coordinate = f"{get_column_letter(column_number)}{row_number + 1}"
                populated.append(f"{coordinate}={_safe_scalar(value)}")
            if not populated:
                continue
            line = " | ".join(populated) + "\n"
            remaining = MAX_SPREADSHEET_LLM_CHARS - used
            if remaining <= 0:
                break
            if len(line) > remaining:
                blocks.append(line[:remaining])
                used += remaining
                break
            blocks.append(line)
            used += len(line)
    if used >= MAX_SPREADSHEET_LLM_CHARS:
        blocks.append("\n[Cell structure truncated at configured limit]\n")
    return "".join(blocks)


def _structural_text(file_path: str) -> str:
    if Path(file_path).suffix.lower() == ".xls":
        return _structural_text_legacy(file_path)
    return _structural_text_modern(file_path)


def _visual_manifest_for_llm(
    visual_manifest: Mapping[str, Any],
) -> dict[str, Any]:
    """Drop exact point arrays before sending the deterministic manifest."""

    available = list(visual_manifest.get("charts", [])) + list(
        visual_manifest.get("images", [])
    )
    selected = available[: max(0, MAX_VISUALS_FOR_LLM)]
    result: dict[str, Any] = {
        "charts": [],
        "images": [],
        "truncated": len(selected) < len(available),
        "total_visuals": len(available),
    }
    # Reserve most of the overall input budget for cell structure while still
    # preserving a useful, bounded visual manifest.
    character_budget = max(256, MAX_SPREADSHEET_LLM_CHARS // 3)

    def bounded(value: Any, limit: int = 500) -> Any:
        if isinstance(value, str):
            return value[:limit]
        return value

    def fits() -> bool:
        return len(json.dumps(result, ensure_ascii=False, default=str)) <= character_budget

    for visual in selected:
        if visual.get("visual_type") == "chart":
            chart_entry = {
                key: bounded(visual.get(key))
                for key in (
                    "visual_id",
                    "sheet_name",
                    "chart_type",
                    "title",
                    "location",
                    "x_axis_title",
                    "y_axis_title",
                )
                if visual.get(key) not in (None, "")
            } | {"series": []}
            result["charts"].append(chart_entry)
            if not fits():
                result["charts"].pop()
                result["truncated"] = True
                break

            for item in visual.get("series", []):
                series_entry = {
                    key: bounded(item.get(key))
                    for key in (
                        "series_index",
                        "series_name",
                        "category_range",
                        "value_range",
                        "x_range",
                        "y_range",
                    )
                    if item.get(key) is not None
                } | {"datapoint_count": len(item.get("datapoints", []))}
                chart_entry["series"].append(series_entry)
                if not fits():
                    chart_entry["series"].pop()
                    result["truncated"] = True
                    break
        else:
            image_entry = {
                key: bounded(visual.get(key))
                for key in (
                    "visual_id",
                    "sheet_name",
                    "location",
                    "format",
                    "width",
                    "height",
                )
                if visual.get(key) is not None
            }
            result["images"].append(image_entry)
            if not fits():
                result["images"].pop()
                result["truncated"] = True
                break
    return result


def _string(value: Any, *, max_length: int = 4000) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip()[:max_length]


def validate_spreadsheet_analysis(
    analysis: Any,
    visual_manifest: Mapping[str, Any],
) -> dict[str, Any]:
    """Allow-list and normalize the LLM response before it is consumed."""

    if not isinstance(analysis, Mapping):
        return _empty_analysis()

    valid_sheet_names = {
        item.get("sheet_name")
        for item in visual_manifest.get("sheets", [])
        if isinstance(item, Mapping) and isinstance(item.get("sheet_name"), str)
    }
    valid_visual_ids = {
        item.get("visual_id")
        for key in ("charts", "images")
        for item in visual_manifest.get(key, [])
        if isinstance(item, Mapping) and isinstance(item.get("visual_id"), str)
    }

    result = _empty_analysis()
    result["spreadsheet_summary"] = _string(
        analysis.get("spreadsheet_summary")
    )

    raw_tables = analysis.get("tables")
    seen_tables: set[tuple[Any, ...]] = set()
    if isinstance(raw_tables, list):
        for raw_table in raw_tables[: max(0, MAX_TABLES_FROM_LLM)]:
            if not isinstance(raw_table, Mapping):
                continue
            sheet_name = _string(raw_table.get("sheet_name"), max_length=255)
            if not sheet_name and len(valid_sheet_names) == 1:
                # Backward-compatible for older single-sheet analysis responses.
                sheet_name = next(iter(valid_sheet_names))
            if sheet_name not in valid_sheet_names:
                logger.warning("Ignoring LLM table with an invalid sheet reference")
                continue
            cell_range = _string(raw_table.get("cell_range"), max_length=100)
            try:
                _local_table_range(cell_range)
            except (TypeError, ValueError):
                logger.warning("Ignoring LLM table with an invalid cell range")
                continue
            raw_headers = raw_table.get("column_headers")
            headers = (
                [
                    _string(header, max_length=255)
                    for header in raw_headers[: max(0, MAX_HEADERS_PER_TABLE)]
                ]
                if isinstance(raw_headers, list)
                else []
            )
            table = {
                "sheet_name": sheet_name,
                "table_name": _string(
                    raw_table.get("table_name"), max_length=500
                )
                or "Unknown Table",
                "cell_range": cell_range,
                "column_headers": headers,
                "description": _string(raw_table.get("description")),
            }
            identity = (
                table["sheet_name"],
                table["cell_range"].upper(),
                tuple(table["column_headers"]),
            )
            if identity not in seen_tables:
                seen_tables.add(identity)
                result["tables"].append(table)

    raw_semantics = analysis.get("visual_semantics")
    seen_visuals: set[str] = set()
    if isinstance(raw_semantics, list):
        for raw_semantic in raw_semantics[: max(0, MAX_VISUALS_FOR_LLM)]:
            if not isinstance(raw_semantic, Mapping):
                continue
            visual_id = _string(raw_semantic.get("visual_id"), max_length=500)
            if visual_id not in valid_visual_ids or visual_id in seen_visuals:
                if visual_id and visual_id not in valid_visual_ids:
                    logger.warning("Ignoring LLM semantics for an unknown visual")
                continue
            seen_visuals.add(visual_id)
            result["visual_semantics"].append(
                {
                    "visual_id": visual_id,
                    "name": _string(raw_semantic.get("name"), max_length=500),
                    "description": _string(raw_semantic.get("description")),
                    "x_axis_semantic": _string(
                        raw_semantic.get("x_axis_semantic"), max_length=500
                    ),
                    "y_axis_semantic": _string(
                        raw_semantic.get("y_axis_semantic"), max_length=500
                    ),
                    "unit": _string(raw_semantic.get("unit"), max_length=255),
                }
            )

    findings = analysis.get("key_findings")
    seen_findings: set[str] = set()
    if isinstance(findings, list):
        for finding in findings[: max(0, MAX_FINDINGS_FROM_LLM)]:
            normalized = _string(finding)
            identity = normalized.casefold()
            if normalized and identity not in seen_findings:
                seen_findings.add(identity)
                result["key_findings"].append(normalized)
    return result


def analyze_spreadsheet_with_llm(
    file_path: str,
    visual_manifest: Mapping[str, Any] | Any | None = None,
    *,
    client: OpenAI | Any | None = None,
) -> dict[str, Any]:
    """Ask the LLM for table boundaries and semantics, never exact values."""

    # Compatibility with the former ``(file_path, client)`` positional shape.
    if visual_manifest is not None and not isinstance(visual_manifest, Mapping):
        if client is None:
            client = visual_manifest
        visual_manifest = None

    manifest = (
        dict(visual_manifest)
        if isinstance(visual_manifest, Mapping)
        else scan_workbook(file_path)
    )
    bounded_manifest = _visual_manifest_for_llm(manifest)
    manifest_json = json.dumps(
        bounded_manifest,
        ensure_ascii=False,
        default=str,
    )
    structural_text = _structural_text(file_path)
    remaining_cell_chars = max(
        0,
        MAX_SPREADSHEET_LLM_CHARS - len(manifest_json),
    )
    if len(structural_text) > remaining_cell_chars:
        marker = "\n[Cell structure truncated at configured limit]\n"
        content_chars = max(0, remaining_cell_chars - len(marker))
        structural_text = structural_text[:content_chars] + (
            marker if remaining_cell_chars >= len(marker) else ""
        )
    prompt = f"""
Analyze the workbook structure below. Deterministic Python extraction is
authoritative for chart types, locations, source ranges, and numeric values.
Do not recreate, guess, or modify chart datapoints. Your job is to identify
logical tables and describe the meaning and intent of existing visual IDs.

Return one valid JSON object and no markdown. Use exactly this schema:
{{
  "spreadsheet_summary": "brief workbook overview",
  "tables": [
    {{
      "sheet_name": "an existing worksheet name",
      "table_name": "logical table name",
      "cell_range": "A1:D20",
      "column_headers": ["Header 1", "Header 2"],
      "description": "what the table represents"
    }}
  ],
  "visual_semantics": [
    {{
      "visual_id": "copy an existing visual_id exactly",
      "name": "human-readable name",
      "description": "what the visual communicates",
      "x_axis_semantic": "meaning of x/category axis",
      "y_axis_semantic": "meaning of y/value axis",
      "unit": "unit or empty string"
    }}
  ],
  "key_findings": ["important supported finding"]
}}

Rules:
- Every table must include the correct existing sheet_name.
- visual_id values must be copied from the deterministic manifest.
- Do not add visual objects that are absent from the manifest.
- Do not return chart type, location, source ranges, series data, or datapoints.

DETERMINISTIC VISUAL MANIFEST:
{manifest_json}

BOUNDED CELL STRUCTURE:
{structural_text}
""".strip()

    try:
        active_client = client or _configured_openai_client()
        response = active_client.chat.completions.create(
            model=SPREADSHEET_LLM_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You analyze spreadsheet structure and semantics and "
                        "output strictly valid JSON."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            response_format={"type": "json_object"},
            temperature=0.0,
            max_tokens=MAX_LLM_OUTPUT_TOKENS,
        )
        raw_content = response.choices[0].message.content
        parsed = json.loads(raw_content or "")
    except Exception:
        logger.warning("Spreadsheet LLM analysis failed")
        return _empty_analysis()
    return validate_spreadsheet_analysis(parsed, manifest)


def _image_bytes(
    file_path: str,
    visual: Mapping[str, Any],
) -> tuple[bytes, str]:
    """Read one embedded raster exactly once from a freshly opened workbook."""

    extension = Path(file_path).suffix.lower()
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=False,
        read_only=False,
        keep_vba=extension == ".xlsm",
    )
    try:
        sheet_name = str(visual.get("sheet_name") or "")
        visual_id = str(visual.get("visual_id") or "")
        try:
            image_index = int(visual_id.rsplit(":image:", 1)[1])
        except (IndexError, TypeError, ValueError) as exc:
            raise ValueError(f"Invalid image visual_id: {visual_id!r}") from exc
        images = _safe_worksheet_images(workbook[sheet_name])
        image = images[image_index]
        raw = image._data()  # Private but the only loaded-image byte boundary.
        if len(raw) > settings.max_archive_uncompressed_mb * 1024 * 1024:
            raise ValueError("Embedded image exceeds the configured upload limit")
        if raw.startswith(b"\x89PNG\r\n\x1a\n"):
            mime_type = "image/png"
        elif raw.startswith(b"\xff\xd8\xff"):
            mime_type = "image/jpeg"
        elif raw.startswith((b"GIF87a", b"GIF89a")):
            mime_type = "image/gif"
        else:
            # openpyxl currently returns original GIF/JPEG/PNG bytes and
            # converts every other raster format to PNG.  Be conservative if a
            # future version returns something else.
            mime_type = "application/octet-stream"
        return raw, mime_type
    finally:
        _safe_close(workbook)


def _validate_visual_analysis(
    payload: Any,
    visual_id: str,
) -> dict[str, Any]:
    if not isinstance(payload, Mapping):
        return {}
    if _string(payload.get("visual_id"), max_length=500) != visual_id:
        return {}

    visible_text = payload.get("visible_text")
    observations = payload.get("key_observations")
    confidence = payload.get("confidence")
    try:
        confidence_value = float(confidence)
        if not 0.0 <= confidence_value <= 1.0:
            confidence_value = None
    except (TypeError, ValueError):
        confidence_value = None

    return {
        "visual_id": visual_id,
        "visual_type": _string(payload.get("visual_type"), max_length=255),
        "name": _string(payload.get("name"), max_length=500),
        "description": _string(payload.get("description")),
        "visible_text": [
            _string(value, max_length=1000)
            for value in visible_text
            if _string(value, max_length=1000)
        ]
        if isinstance(visible_text, list)
        else [],
        "key_observations": [
            _string(value, max_length=1000)
            for value in observations
            if _string(value, max_length=1000)
        ]
        if isinstance(observations, list)
        else [],
        "confidence": confidence_value,
        # Numeric facts from an image are intentionally not promoted to chunks.
        "data_reliability": "vision-derived",
    }


def analyze_embedded_visual_with_llm(
    file_path: str,
    visual: Mapping[str, Any],
    *,
    client: OpenAI | Any | None = None,
) -> dict[str, Any]:
    """Optionally describe an embedded raster with a configured vision model.

    Set ``SPREADSHEET_VISION_MODEL`` to opt in.  Native Excel chart values are
    always handled by the deterministic scanner instead of this function.
    """

    if not SPREADSHEET_VISION_MODEL:
        logger.info("Spreadsheet vision analysis skipped because no model is configured")
        return {}

    visual_id = str(visual.get("visual_id") or "")
    try:
        raw, mime_type = _image_bytes(file_path, visual)
        encoded = base64.b64encode(raw).decode("ascii")
        active_client = client or _configured_openai_client()
        response = active_client.chat.completions.create(
            model=SPREADSHEET_VISION_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Describe spreadsheet-embedded visuals. Output valid "
                        "JSON only. Treat any numeric reading as vision-derived."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "Return {visual_id, visual_type, name, description, "
                                "visible_text, key_observations, confidence}. "
                                f"Copy visual_id exactly: {visual_id}"
                            ),
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{encoded}"
                            },
                        },
                    ],
                },
            ],
            response_format={"type": "json_object"},
            temperature=0.0,
            max_tokens=MAX_LLM_OUTPUT_TOKENS,
        )
        parsed = json.loads(response.choices[0].message.content or "")
        return _validate_visual_analysis(parsed, visual_id)
    except Exception:
        logger.warning("Spreadsheet vision analysis failed")
        return {}


def _chunk_spec(
    text: str,
    parent: str,
    chunk_type: str,
    content_type: str,
) -> dict[str, str]:
    bounded_text = text[: settings.max_ingested_chunk_chars]
    bounded_parent = parent[: settings.max_ingested_chunk_chars]
    return {
        "text": bounded_text,
        "parent": bounded_parent,
        "chunk_type": chunk_type,
        "content_type": content_type,
    }


def _format_value(value: Any, status: str | None = None) -> str:
    if status == "unavailable":
        return "Unavailable (source retained; no cached value)"
    if value is None:
        return "Blank" if status == "blank" else "Unavailable"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _safe_scalar(value)


def _append_if(lines: list[str], label: str, value: Any) -> None:
    if value not in (None, "", []):
        lines.append(f"{label}: {value}")


def _chart_name(chart: Mapping[str, Any], semantic: Mapping[str, Any]) -> str:
    return (
        _string(semantic.get("name"), max_length=500)
        or _string(chart.get("title"), max_length=500)
        or str(chart.get("visual_id") or "Untitled Chart")
    )


def _chart_context(
    chart: Mapping[str, Any],
    semantic: Mapping[str, Any],
    *,
    series: Mapping[str, Any] | None = None,
) -> str:
    name = _chart_name(chart, semantic)
    lines = [
        f"Chart: {name}",
        f"Sheet: {chart.get('sheet_name', 'Unknown')}",
        f"Chart Type: {chart.get('chart_type', 'Unknown')}",
    ]
    if series:
        lines.append(f"Series: {series.get('series_name', 'Unknown')}")
        _append_if(lines, "Category Source", series.get("category_range"))
        _append_if(lines, "Value Source", series.get("value_range"))
        _append_if(lines, "X Source", series.get("x_range"))
        _append_if(lines, "Y Source", series.get("y_range"))
    _append_if(
        lines,
        "X Axis",
        semantic.get("x_axis_semantic")
        or (series or {}).get("x_axis_title")
        or chart.get("x_axis_title"),
    )
    _append_if(
        lines,
        "Y Axis",
        semantic.get("y_axis_semantic")
        or (series or {}).get("y_axis_title")
        or chart.get("y_axis_title"),
    )
    _append_if(lines, "Unit", semantic.get("unit"))
    _append_if(lines, "Description", semantic.get("description"))
    return "\n".join(lines)


def _chart_chunks(
    chart: Mapping[str, Any],
    semantic: Mapping[str, Any],
) -> list[dict[str, str]]:
    name = _chart_name(chart, semantic)
    series_items = [
        item for item in chart.get("series", []) if isinstance(item, Mapping)
    ]
    metadata_lines = [
        f"[Chart: {name}]",
        f"Visual ID: {chart.get('visual_id', '')}",
        f"Sheet: {chart.get('sheet_name', 'Unknown')}",
        f"Chart Type: {chart.get('chart_type', 'Unknown')}",
        f"Location: {chart.get('location', 'Unknown')}",
    ]
    _append_if(
        metadata_lines,
        "Series",
        ", ".join(
            str(item.get("series_name") or "Unknown") for item in series_items
        ),
    )
    for item in series_items:
        series_label = str(item.get("series_name") or "Unknown")
        _append_if(
            metadata_lines,
            f"{series_label} X Axis",
            item.get("x_axis_title"),
        )
        _append_if(
            metadata_lines,
            f"{series_label} Y Axis",
            item.get("y_axis_title"),
        )
        _append_if(
            metadata_lines,
            f"{series_label} Category Source",
            item.get("category_range"),
        )
        _append_if(
            metadata_lines,
            f"{series_label} Value Source",
            item.get("value_range"),
        )
        _append_if(
            metadata_lines,
            f"{series_label} X Source",
            item.get("x_range"),
        )
        _append_if(
            metadata_lines,
            f"{series_label} Y Source",
            item.get("y_range"),
        )
    _append_if(
        metadata_lines,
        "X Axis",
        semantic.get("x_axis_semantic") or chart.get("x_axis_title"),
    )
    _append_if(
        metadata_lines,
        "Y Axis",
        semantic.get("y_axis_semantic") or chart.get("y_axis_title"),
    )
    _append_if(metadata_lines, "Unit", semantic.get("unit"))
    _append_if(metadata_lines, "Description", semantic.get("description"))
    metadata_text = "\n".join(metadata_lines)
    chunks = [
        _chunk_spec(
            metadata_text,
            metadata_text,
            "chart_metadata",
            "metadata",
        )
    ]

    seen_points: set[tuple[Any, ...]] = set()
    for series_position, series in enumerate(series_items):
        series_name = str(series.get("series_name") or f"Series {series_position + 1}")
        series_index = series.get("series_index", series_position)
        parent = _chart_context(chart, semantic, series=series)
        is_scatter = "x_range" in series or "y_range" in series

        for point in series.get("datapoints", []):
            if not isinstance(point, Mapping):
                continue
            if is_scatter:
                x = point.get("x")
                y = point.get("y")
                x_status = str(point.get("x_status") or "")
                y_status = str(point.get("y_status") or "")
                identity = (
                    chart.get("visual_id"),
                    series_index,
                    _safe_scalar(x),
                    _safe_scalar(y),
                )
                point_lines = [
                    f"[Chart: {name}]",
                    f"Sheet: {chart.get('sheet_name', 'Unknown')}",
                    f"Chart Type: {chart.get('chart_type', 'Unknown')}",
                    f"Series: {series_name}",
                    f"X: {_format_value(x, x_status)}",
                    f"Y: {_format_value(y, y_status)}",
                ]
                if x_status == "unavailable":
                    _append_if(point_lines, "X Source", series.get("x_range"))
                if y_status == "unavailable":
                    _append_if(point_lines, "Y Source", series.get("y_range"))
            else:
                category = point.get("category")
                value = point.get("value")
                category_status = str(point.get("category_status") or "")
                value_status = str(point.get("value_status") or "")
                identity = (
                    chart.get("visual_id"),
                    series_index,
                    _safe_scalar(category),
                    _safe_scalar(value),
                )
                point_lines = [
                    f"[Chart: {name}]",
                    f"Sheet: {chart.get('sheet_name', 'Unknown')}",
                    f"Chart Type: {chart.get('chart_type', 'Unknown')}",
                    f"Series: {series_name}",
                    f"Category: {_format_value(category, category_status)}",
                    f"Value: {_format_value(value, value_status)}",
                ]
                if category_status == "unavailable":
                    _append_if(
                        point_lines,
                        "Category Source",
                        series.get("category_range"),
                    )
                if value_status == "unavailable":
                    _append_if(
                        point_lines,
                        "Value Source",
                        series.get("value_range"),
                    )

            if identity in seen_points:
                continue
            seen_points.add(identity)
            _append_if(
                point_lines,
                "X Axis Meaning",
                semantic.get("x_axis_semantic")
                or series.get("x_axis_title")
                or chart.get("x_axis_title"),
            )
            _append_if(
                point_lines,
                "Y Axis Meaning",
                semantic.get("y_axis_semantic")
                or series.get("y_axis_title")
                or chart.get("y_axis_title"),
            )
            _append_if(point_lines, "Unit", semantic.get("unit"))
            chunks.append(
                _chunk_spec(
                    "\n".join(point_lines),
                    parent,
                    "chart_datapoint",
                    "metadata",
                )
            )
    return chunks


def _visual_chunk(
    visual: Mapping[str, Any],
    analysis: Mapping[str, Any] | None = None,
) -> dict[str, str]:
    semantics = analysis or {}
    name = _string(semantics.get("name"), max_length=500) or str(
        visual.get("visual_id") or "Embedded Visual"
    )
    lines = [
        f"[Visual: {name}]",
        f"Visual ID: {visual.get('visual_id', '')}",
        f"Sheet: {visual.get('sheet_name', 'Unknown')}",
        f"Location: {visual.get('location', 'Unknown')}",
    ]
    _append_if(lines, "Format", visual.get("format"))
    if visual.get("width") is not None and visual.get("height") is not None:
        lines.append(f"Dimensions: {visual['width']} x {visual['height']}")
    _append_if(lines, "Visual Type", semantics.get("visual_type"))
    _append_if(lines, "Description", semantics.get("description"))
    for visible in semantics.get("visible_text", []) or []:
        lines.append(f"Visible Text: {visible}")
    for observation in semantics.get("key_observations", []) or []:
        lines.append(f"Observation: {observation}")
    _append_if(lines, "Vision Confidence", semantics.get("confidence"))
    if semantics.get("data_reliability") == "vision-derived":
        lines.append("Reliability: vision-derived semantics")
    text = "\n".join(lines)
    return _chunk_spec(text, text, "visual_metadata", "metadata")


def _fallback_row_chunks(file_path: str) -> list[dict[str, str]]:
    chunks: list[dict[str, str]] = []
    try:
        excel_file = pd.ExcelFile(file_path)
        for sheet_name in excel_file.sheet_names:
            frame = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            for row_index, row in frame.iterrows():
                values = [
                    _safe_scalar(value)
                    for value in row.values
                    if pd.notna(value)
                ]
                if not values:
                    continue
                row_text = (
                    f"Sheet: {sheet_name}\n"
                    f"Row: {row_index + 1}\n"
                    + " | ".join(values)
                )
                chunks.append(
                    _chunk_spec(
                        row_text,
                        row_text,
                        "tabular_record",
                        "tabular",
                    )
                )
    except Exception:
        logger.warning("Spreadsheet raw-row fallback failed")
    return chunks


def _call_visual_analyzer(
    analyzer: Callable[..., Mapping[str, Any]],
    file_path: str,
    visual: Mapping[str, Any],
    client: Any,
) -> dict[str, Any]:
    try:
        analyzer_signature = inspect.signature(analyzer)
        analyzer_signature.bind(file_path, visual, client=client)
    except (TypeError, ValueError):
        result = analyzer(file_path, visual)
    else:
        result = analyzer(file_path, visual, client=client)
    return _validate_visual_analysis(result, str(visual.get("visual_id") or ""))


def build_spreadsheet_chunk_specs(
    file_path: str,
    filename: str,
    *,
    client: OpenAI | Any | None = None,
    analysis: Mapping[str, Any] | None = None,
    visual_analyzer: Callable[..., Mapping[str, Any]] | None = None,
) -> list[dict[str, str]]:
    """Build normalized spreadsheet chunks without embedding or persistence."""

    try:
        manifest = scan_workbook(file_path, filename=filename)
    except Exception:
        logger.warning("Deterministic spreadsheet scan failed")
        manifest = {
            "sheets": _legacy_sheet_manifest(file_path),
            "charts": [],
            "images": [],
        }

    if analysis is None:
        normalized_analysis = analyze_spreadsheet_with_llm(
            file_path,
            manifest,
            client=client,
        )
    else:
        normalized_analysis = validate_spreadsheet_analysis(analysis, manifest)

    chunks: list[dict[str, str]] = []
    table_chunks_added = 0
    for table in normalized_analysis["tables"]:
        try:
            row_chunks = chunk_table_deterministically(file_path, table)
            for row_text in row_chunks:
                chunks.append(
                    _chunk_spec(
                        row_text,
                        row_text,
                        "tabular_record",
                        "tabular",
                    )
                )
                table_chunks_added += 1
        except Exception:
            logger.warning("Spreadsheet table chunk generation failed")

    if not normalized_analysis["tables"] or table_chunks_added == 0:
        logger.warning("Spreadsheet has no usable logical tables; using row fallback")
        chunks.extend(_fallback_row_chunks(file_path))

    seen_metadata: set[str] = set()
    summary = normalized_analysis["spreadsheet_summary"]
    if summary:
        seen_metadata.add(summary.casefold())
        summary_text = f"Workbook Summary:\n{summary}"
        chunks.append(
            _chunk_spec(
                summary_text,
                summary_text,
                "workbook_metadata",
                "metadata",
            )
        )

    for finding in normalized_analysis["key_findings"]:
        identity = finding.casefold()
        if identity in seen_metadata:
            continue
        seen_metadata.add(identity)
        finding_text = f"Workbook Finding:\n{finding}"
        chunks.append(
            _chunk_spec(
                finding_text,
                finding_text,
                "workbook_metadata",
                "metadata",
            )
        )

    semantics_by_id = {
        semantic["visual_id"]: semantic
        for semantic in normalized_analysis["visual_semantics"]
    }
    for chart in manifest.get("charts", []):
        try:
            chunks.extend(
                _chart_chunks(
                    chart,
                    semantics_by_id.get(chart.get("visual_id"), {}),
                )
            )
        except Exception:
            logger.warning("Spreadsheet chart chunk generation failed")

    active_visual_analyzer = visual_analyzer
    if active_visual_analyzer is None and SPREADSHEET_VISION_MODEL:
        active_visual_analyzer = analyze_embedded_visual_with_llm

    for visual in manifest.get("images", []):
        visual_analysis: dict[str, Any] = dict(
            semantics_by_id.get(visual.get("visual_id"), {})
        )
        if active_visual_analyzer is not None:
            try:
                vision_analysis = _call_visual_analyzer(
                    active_visual_analyzer,
                    file_path,
                    visual,
                    client,
                )
                visual_analysis.update(
                    {
                        key: value
                        for key, value in vision_analysis.items()
                        if value not in (None, "", [])
                    }
                )
            except Exception:
                logger.warning("Spreadsheet visual analysis failed")
        chunks.append(_visual_chunk(visual, visual_analysis))

    return chunks


__all__ = [
    "MAX_SPREADSHEET_LLM_CHARS",
    "MAX_VISUALS_FOR_LLM",
    "analyze_embedded_visual_with_llm",
    "analyze_spreadsheet_with_llm",
    "build_spreadsheet_chunk_specs",
    "chunk_table_deterministically",
    "scan_workbook",
    "validate_spreadsheet_analysis",
]
