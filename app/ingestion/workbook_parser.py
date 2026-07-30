import re
import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
import openpyxl
from .models import SheetData, WorkbookData, TableRegion


def extract_workbook(file_source: Union[bytes, str, Path], filename: str = "uploaded_spreadsheet.xlsx") -> WorkbookData:
    """Phase 1: Reads an Excel file (from raw bytes or path) and preserves structural metadata,
    merged cells, hidden regions, comments, and formulas.
    """
    if isinstance(file_source, bytes):
        source_values = io.BytesIO(file_source)
        source_formulas = io.BytesIO(file_source)
        display_filename = filename
    else:
        path = Path(file_source)
        source_values = path
        source_formulas = path
        display_filename = path.name

    # Load workbook twice: data_only=True for evaluated values, data_only=False for raw formulas.
    wb_values = openpyxl.load_workbook(source_values, data_only=True)
    wb_formulas = openpyxl.load_workbook(source_formulas, data_only=False)

    sheets_data: List[SheetData] = []

    for sheet_name in wb_values.sheetnames:
        ws_val = wb_values[sheet_name]
        ws_form = wb_formulas[sheet_name]

        # Sheet visibility
        is_hidden = ws_val.sheet_state != "visible"

        # Merged ranges
        merged_ranges = [str(r) for r in ws_val.merged_cells.ranges]

        # Hidden rows and columns
        hidden_rows = [
            r_idx
            for r_idx, r_dim in ws_val.row_dimensions.items()
            if r_dim.hidden
        ]
        hidden_cols = [
            c_letter
            for c_letter, c_dim in ws_val.column_dimensions.items()
            if c_dim.hidden
        ]

        # Charts presence
        has_charts = len(getattr(ws_val, "_charts", [])) > 0

        # Extract comments and grid data
        comments: Dict[str, str] = {}
        grid_values: List[List[Any]] = []
        grid_formulas: List[List[Optional[str]]] = []

        for row_val, row_form in zip(ws_val.iter_rows(), ws_form.iter_rows()):
            val_row_data = []
            form_row_data = []

            for cell_val, cell_form in zip(row_val, row_form):
                # Save comments
                if cell_val.comment and cell_val.comment.text:
                    comments[cell_val.coordinate] = cell_val.comment.text.strip()
                
                # Values and formulas
                val_row_data.append(cell_val.value)

                # Store formula string if cell contains a formula, else None
                if (
                    isinstance(cell_form.value, str)
                    and cell_form.value.startswith("=")
                ):
                    form_row_data.append(cell_form.value)
                else:
                    form_row_data.append(None)

            grid_values.append(val_row_data)
            grid_formulas.append(form_row_data)

        sheets_data.append(
            SheetData(
                name=sheet_name,
                is_hidden=is_hidden,
                merged_ranges=merged_ranges,
                hidden_rows=hidden_rows,
                hidden_cols=hidden_cols,
                has_charts=has_charts,
                comments=comments,
                grid_values=grid_values,
                grid_formulas=grid_formulas,
            )
        )

    return WorkbookData(filename=display_filename, sheets=sheets_data)


def _parse_cell_ref(ref: str) -> Tuple[int, int]:
    """Converts Excel cell ref like 'C5' to 0-based (row, col) tuple (4, 2)."""
    match = re.match(r"([A-Z]+)([0-9]+)", ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {ref}")
    col_str, row_str = match.groups()
    
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char) - ord('A') + 1)
    return int(row_str) - 1, col - 1


def _parse_range_ref(range_str: str) -> Tuple[int, int, int, int]:
    """Converts 'A1:C3' to 0-based bounding box (min_r, min_c, max_r, max_c)."""
    if ":" in range_str:
        start_ref, end_ref = range_str.split(":")
    else:
        start_ref = end_ref = range_str
    
    start_r, start_c = _parse_cell_ref(start_ref)
    end_r, end_c = _parse_cell_ref(end_ref)
    return min(start_r, end_r), min(start_c, end_c), max(start_r, end_r), max(start_c, end_c)


def normalize_sheet_grid(
    grid_values: List[List[Any]], 
    merged_ranges: List[str]
) -> List[List[Any]]:
    """Copies top-left value across all cells in merged ranges to eliminate holes."""
    # Deep copy grid values
    normalized = [list(row) for row in grid_values]

    for range_str in merged_ranges:
        try:
            min_r, min_c, max_r, max_c = _parse_range_ref(range_str)
            # Ensure boundaries fit within extracted grid size
            max_r = min(max_r, len(normalized) - 1)
            max_c = min(max_c, len(normalized[0]) - 1 if normalized else 0)

            top_left_val = normalized[min_r][min_c]

            # Forward fill all cells in the merged range
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    normalized[r][c] = top_left_val
        except Exception:
            continue

    return normalized

def detect_bounding_boxes(
    grid: List[List[Any]], 
    row_gap_tolerance: int = 1, 
    col_gap_tolerance: int = 1
) -> List[Tuple[int, int, int, int]]:
    """Detects isolated clusters of non-empty cells on a grid.
    
    row_gap_tolerance: Allows tables with 1 empty row inside them to stay as 1 table.
    col_gap_tolerance: Allows tables with 1 empty column inside them to stay as 1 table.
    """
    if not grid or not grid[0]:
        return []

    num_rows = len(grid)
    num_cols = len(grid[0])

    visited = set()
    boxes: List[Tuple[int, int, int, int]] = []

    def is_non_empty(r: int, c: int) -> bool:
        val = grid[r][c]
        return val is not None and str(val).strip() != ""

    for r in range(num_rows):
        for c in range(num_cols):
            if (r, c) in visited or not is_non_empty(r, c):
                continue

            # BFS / Flood Fill to find all connected non-empty cells
            queue = [(r, c)]
            visited.add((r, c))

            cluster_min_r, cluster_max_r = r, r
            cluster_min_c, cluster_max_c = c, c

            while queue:
                curr_r, curr_c = queue.pop(0)

                cluster_min_r = min(cluster_min_r, curr_r)
                cluster_max_r = max(cluster_max_r, curr_r)
                cluster_min_c = min(cluster_min_c, curr_c)
                cluster_max_c = max(cluster_max_c, curr_c)

                # Check neighbors within tolerance limits
                r_start = max(0, curr_r - row_gap_tolerance - 1)
                r_end = min(num_rows - 1, curr_r + row_gap_tolerance + 1)
                c_start = max(0, curr_c - col_gap_tolerance - 1)
                c_end = min(num_cols - 1, curr_c + col_gap_tolerance + 1)

                for nr in range(r_start, r_end + 1):
                    for nc in range(c_start, c_end + 1):
                        if (nr, nc) not in visited and is_non_empty(nr, nc):
                            visited.add((nr, nc))
                            queue.append((nr, nc))

            # Filter out micro-noise (e.g. single orphan cells)
            boxes.append((cluster_min_r, cluster_min_c, cluster_max_r, cluster_max_c))

    return boxes

def extract_regions_from_sheet(sheet: SheetData) -> List[TableRegion]:
    """Extracts table regions from sheet grid data for pipeline processing."""
    regions = []
    if not sheet.grid_values:
        return regions

    max_row = len(sheet.grid_values)
    max_col = len(sheet.grid_values[0]) if max_row > 0 else 0

    region_id = f"{sheet.name}_region_0"
    regions.append(
        TableRegion(
            sheet_name=sheet.name,
            region_id=region_id,
            bounding_box=(0, 0, max_row, max_col),
            grid_values=sheet.grid_values,
            formulas=sheet.grid_formulas,
            comments=sheet.comments
        )
    )
    
    return regions