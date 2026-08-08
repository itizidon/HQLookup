from __future__ import annotations

import csv
import io
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.ingestion.workbook_models import (
    WorkbookFrame,
    WorksheetFrame,
    CellData,
)


# ==========================================================
# Public API
# ==========================================================

def load_workbook_frame(
    file_bytes: bytes,
    filename: str,
) -> WorkbookFrame:
    """
    Loads an XLSX/XLS/CSV into a WorkbookFrame.

    This function ONLY parses the workbook.

    It does NOT:
        - detect tables
        - call an LLM
        - build chunks
        - generate embeddings
    """

    extension = os.path.splitext(filename)[1].lower()

    if extension == ".xls":
        raise NotImplementedError(
            ".xls files are not yet supported. Please convert to .xlsx."
        )

    if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_excel(file_bytes, filename)

    if extension == ".csv":
        return _load_csv(file_bytes, filename)

    raise ValueError(f"Unsupported workbook type: {extension}")


# ==========================================================
# Excel Loader
# ==========================================================

def _load_excel(
    file_bytes: bytes,
    filename: str,
) -> WorkbookFrame:

    workbook = load_workbook(
        filename=io.BytesIO(file_bytes),
        data_only=False,
    )

    workbook_frame = WorkbookFrame(filename=filename)

    workbook_frame.workbook_properties = {
        "sheet_count": len(workbook.sheetnames),
        "active_sheet": workbook.active.title if workbook.active else None,
        "defined_names": list(workbook.defined_names.keys()),
        "creator": getattr(workbook.properties, "creator", None),
        "created": getattr(workbook.properties, "created", None),
        "modified": getattr(workbook.properties, "modified", None),
    }

    for index, worksheet in enumerate(workbook.worksheets):

        workbook_frame.sheets.append(
            _build_sheet(
                worksheet,
                sheet_index=index,
            )
    )

    return workbook_frame


# ==========================================================
# CSV Loader
# ==========================================================

def _load_csv(
    file_bytes: bytes,
    filename: str,
) -> WorkbookFrame:

    workbook = WorkbookFrame(filename=filename)

    worksheet = WorksheetFrame(
        name="Sheet1",
        hidden=False,
    )

    decoded = file_bytes.decode("utf-8", errors="ignore")

    sample = decoded[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(
        io.StringIO(decoded),
        dialect=dialect,
    )

    for row_number, row in enumerate(reader, start=1):

        row_cells = []

        for column_number, value in enumerate(row, start=1):

            coordinate = _coordinate(row_number, column_number)

            row_cells.append(
                CellData(
                    row=row_number,
                    column=column_number,
                    coordinate=coordinate,
                    value=value,
                    data_type="string",
                )
            )

        worksheet.cells.append(row_cells)

    worksheet.max_rows = len(worksheet.cells)

    worksheet.max_columns = max(
        (len(r) for r in worksheet.cells),
        default=0,
    )

    workbook.sheets.append(worksheet)

    workbook.workbook_properties = {
        "sheet_count": 1
    }

    return workbook


# ==========================================================
# Worksheet Builder
# ==========================================================

def _build_sheet(
    ws,
    sheet_index: int,
) -> WorksheetFrame:

    sheet = WorksheetFrame(
        name=ws.title,
        hidden=(ws.sheet_state != "visible"),
        sheet_index=sheet_index,
    )

    sheet.max_rows = ws.max_row
    sheet.max_columns = ws.max_column

    if ws.freeze_panes:
        sheet.frozen_panes = str(ws.freeze_panes)

    sheet.merged_ranges = [
        str(r)
        for r in ws.merged_cells.ranges
    ]

    sheet.excel_tables = [
        {
            "name": table.name,
            "range": table.ref,
        }
        for table in ws.tables.values()
    ]

    merged_lookup = _build_merged_lookup(ws)

    for row_idx, row_dim in ws.row_dimensions.items():

        if row_dim.height is not None:

            sheet.row_heights[row_idx] = row_dim.height

    for col_letter, col_dim in ws.column_dimensions.items():

        if col_dim.width is not None:

            sheet.column_widths[col_letter] = col_dim.width

    for row in ws.iter_rows():

        row_cells = []

        for cell in row:

            row_cells.append(
                _build_cell(
                    cell,
                    merged_lookup,
                )
            )

        sheet.cells.append(row_cells)

    return sheet


# ==========================================================
# Cell Builder
# ==========================================================

def _build_cell(
    cell,
    merged_lookup: set[str],
) -> CellData:

    formula = None

    if cell.data_type == "f":
        formula = cell.value

    fill_color = None

    if isinstance(cell.fill, PatternFill):
        if cell.fill.fill_type:
            fill_color = cell.fill.fgColor.rgb

    return CellData(
        row=cell.row,
        column=cell.column,
        coordinate=cell.coordinate,

        value=cell.value,

        formula=formula,

        data_type=cell.data_type,

        number_format=cell.number_format,

        comment=cell.comment.text if cell.comment else None,

        merged=cell.coordinate in merged_lookup,

        hidden=(
            cell.parent.row_dimensions[cell.row].hidden
            or cell.parent.column_dimensions[cell.column_letter].hidden
        ),

        bold=cell.font.bold,

        italic=cell.font.italic,

        fill_color=fill_color,

        font_color=cell.font.color.rgb
        if cell.font.color and hasattr(cell.font.color, "rgb")
        else None,

        alignment=cell.alignment.horizontal,
    )


# ==========================================================
# Helpers
# ==========================================================

def _build_merged_lookup(ws) -> set[str]:
    """
    Returns every coordinate inside merged ranges.

    Example

    A1:C3

    returns

    A1
    A2
    A3
    B1
    ...
    """

    merged = set()

    for rng in ws.merged_cells.ranges:

        for row in ws[rng.coord]:

            for cell in row:
                merged.add(cell.coordinate)

    return merged


def _coordinate(
    row: int,
    column: int,
) -> str:

    letters = ""

    while column:

        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters

    return f"{letters}{row}"