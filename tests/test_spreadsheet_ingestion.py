"""Regression tests for deterministic spreadsheet ingestion.

All workbooks are generated in the test that consumes them.  In particular, the
tests do not need Excel, a calculation engine, API credentials, or checked-in
binary fixtures.
"""

from __future__ import annotations

import json
import zipfile
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.data_source import NumData, NumVal, StrData, StrRef, StrVal
from openpyxl.chart.text import Text
from openpyxl.chart.title import Title
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PillowImage

from app.services import spreadsheet_ingestion as ingestion


class FakeOpenAIClient:
    """Small OpenAI-compatible client that records prompts and returns JSON."""

    def __init__(self, payload: dict) -> None:
        self.payload = payload
        self.calls: list[dict] = []
        self.chat = SimpleNamespace(
            completions=SimpleNamespace(create=self._create),
        )

    def _create(self, **kwargs):
        self.calls.append(kwargs)
        message = SimpleNamespace(content=json.dumps(self.payload))
        return SimpleNamespace(choices=[SimpleNamespace(message=message)])


class FakeRawOpenAIClient(FakeOpenAIClient):
    """OpenAI-compatible client for malformed/non-JSON response coverage."""

    def __init__(self, raw_content: str) -> None:
        super().__init__({})
        self.raw_content = raw_content

    def _create(self, **kwargs):
        self.calls.append(kwargs)
        message = SimpleNamespace(content=self.raw_content)
        return SimpleNamespace(choices=[SimpleNamespace(message=message)])


def _save(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    return path


def _replace_zip_members(
    path: Path,
    replacements: dict[str, bytes],
    *,
    removed: set[str] | None = None,
) -> None:
    """Rewrite selected XLSX package members for interoperability fixtures."""

    rewritten = BytesIO()
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        rewritten, "w", zipfile.ZIP_DEFLATED
    ) as destination:
        for member in source.infolist():
            if member.filename in replacements or member.filename in (removed or set()):
                continue
            destination.writestr(member, source.read(member.filename))
        for member_name, content in replacements.items():
            destination.writestr(member_name, content)
    path.write_bytes(rewritten.getvalue())


def _line_workbook(
    path: Path,
    *,
    sheet_name: str = "Data",
    multi_series: bool = False,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Month", "Revenue", "Expenses"])
    sheet.append(["Jan", 100, 60])
    sheet.append(["Feb", 120, 75])
    sheet.append(["Mar", 150, 90])

    chart = LineChart()
    chart.title = "Financial Performance" if multi_series else "Monthly Revenue"
    max_col = 3 if multi_series else 2
    chart.add_data(
        Reference(sheet, min_col=2, max_col=max_col, min_row=1, max_row=4),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=4))
    sheet.add_chart(chart, "E2")
    return _save(workbook, path)


def _analysis_for_chart(*, visual_id: str = "Data:chart:0") -> dict:
    return {
        "spreadsheet_summary": "Revenue workbook.",
        "tables": [],
        "visual_semantics": [
            {
                "visual_id": visual_id,
                "name": "Monthly Revenue",
                "description": "Monthly revenue trend.",
                "x_axis_semantic": "Month",
                "y_axis_semantic": "Revenue",
                "unit": "USD",
            }
        ],
        "key_findings": ["Revenue increased each month."],
    }


def _only_chart(scan: dict) -> dict:
    assert len(scan["charts"]) == 1
    return scan["charts"][0]


def _chunks_of_type(chunks: list[dict], chunk_type: str) -> list[dict]:
    return [chunk for chunk in chunks if chunk["chunk_type"] == chunk_type]


def test_basic_table_rows_keep_existing_format(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Revenue"
    sheet.append(["Month", "Revenue"])
    sheet.append(["Jan", 100])
    sheet.append(["Feb", 120])
    path = _save(workbook, tmp_path / "table.xlsx")

    chunks = ingestion.chunk_table_deterministically(
        str(path),
        {
            "sheet_name": "Revenue",
            "table_name": "Monthly Revenue",
            "cell_range": "A1:B3",
            "column_headers": ["Month", "Revenue"],
            "description": "Revenue by month.",
        },
    )

    assert len(chunks) == 2
    assert "[Table: Monthly Revenue]" in chunks[0]
    assert "Headers: Month, Revenue" in chunks[0]
    assert "Row Data: Month: Jan | Revenue: 100" in chunks[0]
    assert "Row Data: Month: Feb | Revenue: 120" in chunks[1]

    specs = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Revenue workbook.",
            "tables": [
                {
                    "sheet_name": "Revenue",
                    "table_name": "Monthly Revenue",
                    "cell_range": "A1:B3",
                    "column_headers": ["Month", "Revenue"],
                    "description": "Revenue by month.",
                }
            ],
            "visual_semantics": [],
            "key_findings": [],
        },
    )
    table_specs = _chunks_of_type(specs, "tabular_record")
    assert len(table_specs) == 2
    assert all(spec["content_type"] == "tabular" for spec in table_specs)
    assert all(spec["parent"] == spec["text"] for spec in table_specs)
    assert all(
        set(spec) == {"text", "parent", "chunk_type", "content_type"}
        for spec in specs
    )


def test_blank_header_does_not_shift_later_column_labels(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Messy"
    sheet.append(["Region", None, "Revenue"])
    sheet.append(["North", "Internal note", 100])
    path = _save(workbook, tmp_path / "blank-header.xlsx")

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Messy table.",
            "tables": [
                {
                    "sheet_name": "Messy",
                    "table_name": "Regional Revenue",
                    "cell_range": "A1:C2",
                    "column_headers": ["Region", "", "Revenue"],
                    "description": "Revenue with an unlabeled notes column.",
                }
            ],
            "visual_semantics": [],
            "key_findings": [],
        },
    )

    rows = _chunks_of_type(chunks, "tabular_record")
    assert len(rows) == 1
    assert (
        "Row Data: Region: North | Internal note | Revenue: 100"
        in rows[0]["text"]
    )


def test_line_chart_scan_and_chunk_generation_preserve_point_alignment(
    tmp_path: Path,
) -> None:
    path = _line_workbook(tmp_path / "line.xlsx")

    chart = _only_chart(ingestion.scan_workbook(str(path)))
    assert chart["visual_id"] == "Data:chart:0"
    assert chart["visual_type"] == "chart"
    assert chart["sheet_name"] == "Data"
    assert chart["chart_type"] == "LineChart"
    assert chart["title"] == "Monthly Revenue"
    assert chart["location"] == "E2"

    series = chart["series"][0]
    assert series["series_name"] == "Revenue"
    assert series["category_range"] == "'Data'!$A$2:$A$4"
    assert series["value_range"] == "'Data'!$B$2:$B$4"
    assert [
        (point["category"], point["value"])
        for point in series["datapoints"]
    ] == [("Jan", 100), ("Feb", 120), ("Mar", 150)]

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis=_analysis_for_chart(),
    )
    metadata = _chunks_of_type(chunks, "chart_metadata")
    datapoints = _chunks_of_type(chunks, "chart_datapoint")
    assert len(metadata) == 1
    assert metadata[0]["content_type"] == "metadata"
    assert "[Chart: Monthly Revenue]" in metadata[0]["text"]
    assert "Chart Type: LineChart" in metadata[0]["text"]
    assert len(datapoints) == 3
    assert all(chunk["content_type"] == "metadata" for chunk in datapoints)

    march = next(chunk for chunk in datapoints if "Category: Mar" in chunk["text"])
    assert "Series: Revenue" in march["text"]
    assert "Value: 150" in march["text"]
    assert "Value: 120" not in march["text"]
    assert "X Axis: Month" in march["parent"]
    assert "Y Axis: Revenue" in march["parent"]
    assert "Unit: USD" in march["parent"]


def test_multiple_series_never_lose_series_identity(tmp_path: Path) -> None:
    path = _line_workbook(tmp_path / "multi.xlsx", multi_series=True)

    chart = _only_chart(ingestion.scan_workbook(str(path)))
    assert [series["series_name"] for series in chart["series"]] == [
        "Revenue",
        "Expenses",
    ]
    revenue, expenses = chart["series"]
    assert [(p["category"], p["value"]) for p in revenue["datapoints"]] == [
        ("Jan", 100),
        ("Feb", 120),
        ("Mar", 150),
    ]
    assert [(p["category"], p["value"]) for p in expenses["datapoints"]] == [
        ("Jan", 60),
        ("Feb", 75),
        ("Mar", 90),
    ]

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis=_analysis_for_chart(visual_id="Data:chart:0"),
    )
    datapoint_texts = [
        chunk["text"] for chunk in _chunks_of_type(chunks, "chart_datapoint")
    ]
    assert len(datapoint_texts) == 6
    assert any(
        "Series: Revenue" in text
        and "Category: Mar" in text
        and "Value: 150" in text
        for text in datapoint_texts
    )
    assert any(
        "Series: Expenses" in text
        and "Category: Mar" in text
        and "Value: 90" in text
        for text in datapoint_texts
    )


def test_scatter_chart_uses_x_and_y_ranges_and_datapoints(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scatter Data"
    sheet.append(["Units", "Revenue"])
    sheet.append([1, 100])
    sheet.append([2, 120])
    sheet.append([3, 150])

    chart = ScatterChart()
    chart.title = "Revenue by Units"
    chart.series.append(
        Series(
            Reference(sheet, min_col=2, min_row=1, max_row=4),
            Reference(sheet, min_col=1, min_row=2, max_row=4),
            title_from_data=True,
        )
    )
    sheet.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / "scatter.xlsx")

    scanned = _only_chart(ingestion.scan_workbook(str(path)))
    assert scanned["chart_type"] == "ScatterChart"
    series = scanned["series"][0]
    assert series["series_name"] == "Revenue"
    assert series["x_range"] == "'Scatter Data'!$A$2:$A$4"
    assert series["y_range"] == "'Scatter Data'!$B$2:$B$4"
    assert [(point["x"], point["y"]) for point in series["datapoints"]] == [
        (1, 100),
        (2, 120),
        (3, 150),
    ]

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Scatter workbook.",
            "tables": [],
            "visual_semantics": [
                {
                    "visual_id": "Scatter Data:chart:0",
                    "name": "Revenue by Units",
                    "description": "Revenue for each units value.",
                    "x_axis_semantic": "Units",
                    "y_axis_semantic": "Revenue",
                    "unit": "USD",
                }
            ],
            "key_findings": [],
        },
    )
    datapoint_chunks = _chunks_of_type(chunks, "chart_datapoint")
    assert len(datapoint_chunks) == 3
    final_point = next(chunk for chunk in datapoint_chunks if "X: 3" in chunk["text"])
    assert "Series: Revenue" in final_point["text"]
    assert "Y: 150" in final_point["text"]
    assert "X Axis: Units" in final_point["parent"]
    assert "Y Axis: Revenue" in final_point["parent"]


@pytest.mark.parametrize(
    ("chart_class", "expected_type"),
    [
        (BarChart, "BarChart"),
        (PieChart, "PieChart"),
        (DoughnutChart, "DoughnutChart"),
        (AreaChart, "AreaChart"),
    ],
)
def test_common_categorical_chart_families_are_scanned(
    tmp_path: Path,
    chart_class,
    expected_type: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "O'Brien"
    sheet.append(["Month", "Revenue"])
    sheet.append(["Jan", 100])
    sheet.append(["Feb", 120])
    chart = chart_class()
    chart.title = "Revenue"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / f"{expected_type}.xlsx")

    scanned = _only_chart(ingestion.scan_workbook(str(path)))
    assert scanned["chart_type"] == expected_type
    assert scanned["series"][0]["category_range"] == "'O''Brien'!$A$2:$A$3"
    assert scanned["series"][0]["value_range"] == "'O''Brien'!$B$2:$B$3"
    assert [(p["category"], p["value"]) for p in scanned["series"][0]["datapoints"]] == [
        ("Jan", 100),
        ("Feb", 120),
    ]


def test_multilevel_categories_remain_one_composite_category_per_value(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Region", "Month", "Revenue"])
    sheet.append(["North", "Jan", 100])
    sheet.append(["North", "Feb", 120])
    sheet.append(["South", "Jan", 90])
    chart = LineChart()
    chart.title = "Regional Revenue"
    chart.add_data(
        Reference(sheet, min_col=3, min_row=1, max_row=4),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=1, max_col=2, min_row=2, max_row=4)
    )
    sheet.add_chart(chart, "E2")
    path = _save(workbook, tmp_path / "multilevel-categories.xlsx")

    series = _only_chart(ingestion.scan_workbook(str(path)))["series"][0]
    assert series["category_range"] == "'Data'!$A$2:$B$4"
    assert [(p["category"], p["value"]) for p in series["datapoints"]] == [
        ("North / Jan", 100),
        ("North / Feb", 120),
        ("South / Jan", 90),
    ]


def test_missing_series_title_gets_stable_fallback(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Untitled"
    sheet.append(["Jan", 100])
    sheet.append(["Feb", 120])
    chart = LineChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=2))
    chart.set_categories(Reference(sheet, min_col=1, min_row=1, max_row=2))
    sheet.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / "untitled-series.xlsx")

    series = _only_chart(ingestion.scan_workbook(str(path)))["series"][0]
    assert series["series_name"] == "Series 1"
    assert [(p["category"], p["value"]) for p in series["datapoints"]] == [
        ("Jan", 100),
        ("Feb", 120),
    ]


def test_resolved_blank_title_cells_override_stale_chart_caches(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Month", "Temporary header", None])
    sheet.append(["Jan", 100, None])
    sheet.append(["Feb", 120, None])
    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    # Both references resolve successfully to exact blank, non-formula cells.
    # Stale OOXML caches must not resurrect old labels.
    sheet["B1"] = None
    chart.ser[0].tx.strRef.strCache = StrData(
        ptCount=1,
        pt=[StrVal(idx=0, v="Stale Series")],
    )
    chart.title = Title(
        tx=Text(
            strRef=StrRef(
                f="'Data'!$C$1",
                strCache=StrData(
                    ptCount=1,
                    pt=[StrVal(idx=0, v="Stale Chart Title")],
                ),
            )
        )
    )
    sheet.add_chart(chart, "E2")
    path = _save(workbook, tmp_path / "stale-titles.xlsx")

    scanned = _only_chart(ingestion.scan_workbook(str(path)))
    assert scanned["title"] == ""
    assert scanned["series"][0]["series_name"] == "Series 1"


def test_chart_on_chartsheet_is_discovered_with_exact_source_data(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["Month", "Revenue"])
    data.append(["Jan", 100])
    data.append(["Feb", 120])
    chart = LineChart()
    chart.title = "Chartsheet Revenue"
    chart.add_data(
        Reference(data, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=3))
    chart_sheet = workbook.create_chartsheet("Revenue Chart")
    chart_sheet.add_chart(chart)
    path = _save(workbook, tmp_path / "chartsheet.xlsx")

    scanned = _only_chart(ingestion.scan_workbook(str(path)))
    assert scanned["visual_id"] == "Revenue Chart:chart:0"
    assert scanned["sheet_name"] == "Revenue Chart"
    assert scanned["title"] == "Chartsheet Revenue"
    assert scanned["series"][0]["category_range"] == "'Data'!$A$2:$A$3"
    assert scanned["series"][0]["value_range"] == "'Data'!$B$2:$B$3"
    assert [(p["category"], p["value"]) for p in scanned["series"][0]["datapoints"]] == [
        ("Jan", 100),
        ("Feb", 120),
    ]


def test_table_extraction_uses_declared_non_first_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Cover"
    first.append(["Do not ingest this as Billing"])
    billing = workbook.create_sheet("Billing")
    billing.append(["Invoice", "Amount"])
    billing.append(["INV-42", 375])
    billing.append(["INV-43", 425])
    chart = LineChart()
    chart.title = "Invoice Amounts"
    chart.add_data(
        Reference(billing, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(billing, min_col=1, min_row=2, max_row=3))
    billing.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / "multiple-sheets.xlsx")

    chunks = ingestion.chunk_table_deterministically(
        str(path),
        {
            "sheet_name": "Billing",
            "table_name": "Invoices",
            "cell_range": "A1:B3",
            "column_headers": ["Invoice", "Amount"],
            "description": "Outstanding invoices.",
        },
    )

    assert len(chunks) == 2
    assert "Invoice: INV-42 | Amount: 375" in chunks[0]
    assert "Do not ingest" not in chunks[0]
    scanned = _only_chart(ingestion.scan_workbook(str(path)))
    assert scanned["visual_id"] == "Billing:chart:0"
    assert scanned["sheet_name"] == "Billing"
    assert scanned["series"][0]["category_range"] == "'Billing'!$A$2:$A$3"
    assert [(p["category"], p["value"]) for p in scanned["series"][0]["datapoints"]] == [
        ("INV-42", 375),
        ("INV-43", 425),
    ]


def test_formula_values_use_chart_cache_when_cell_cache_is_missing(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Forecast"
    sheet.append(["Month", "Forecast"])
    sheet.append(["Jan", "=1+1"])
    sheet.append(["Feb", "=2+2"])
    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    chart.ser[0].val.numRef.numCache = NumData(
        formatCode="General",
        ptCount=2,
        pt=[NumVal(idx=0, v=2), NumVal(idx=1, v=4)],
    )
    sheet.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / "cached-formulas.xlsx")

    series = _only_chart(ingestion.scan_workbook(str(path)))["series"][0]
    assert series["value_range"] == "'Forecast'!$B$2:$B$3"
    assert [point["value"] for point in series["datapoints"]] == [2, 4]
    assert all(point["value"] not in {"=1+1", "=2+2"} for point in series["datapoints"])


def test_exact_cell_values_take_priority_over_stale_chart_cache(tmp_path: Path) -> None:
    path = _line_workbook(tmp_path / "stale-cache.xlsx")

    # Add a deliberately stale chart cache.  Exact referenced workbook cells
    # remain authoritative and therefore must win.
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    chart = workbook["Data"]._charts[0]
    chart.ser[0].val.numRef.numCache = NumData(
        formatCode="General",
        ptCount=3,
        pt=[
            NumVal(idx=0, v=999),
            NumVal(idx=1, v=998),
            NumVal(idx=2, v=997),
        ],
    )
    workbook.save(path)

    series = _only_chart(ingestion.scan_workbook(str(path)))["series"][0]
    assert [point["value"] for point in series["datapoints"]] == [100, 120, 150]


def test_formula_without_any_cache_is_marked_unavailable_not_invented(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Forecast"
    sheet.append(["Month", "Forecast"])
    sheet.append(["Jan", "=1+1"])
    sheet.append(["Feb", "=2+2"])
    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / "uncached-formulas.xlsx")

    series = _only_chart(ingestion.scan_workbook(str(path)))["series"][0]
    assert series["value_range"] == "'Forecast'!$B$2:$B$3"
    assert all(point.get("value") is None for point in series["datapoints"])
    assert "unavailable" in json.dumps(series).lower()
    assert "=1+1" not in json.dumps(series["datapoints"])


def test_broken_series_reference_does_not_discard_valid_series(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Month", "Revenue", "Expenses"])
    sheet.append(["Jan", 100, 60])
    sheet.append(["Feb", 120, 75])
    sheet.append(["Mar", 150, 90])
    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=2, max_col=3, min_row=1, max_row=4),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=4))
    broken_reference = "'Missing Sheet'!$C$2:$C$4"
    chart.ser[1].val.numRef.f = broken_reference
    sheet.add_chart(chart, "E2")
    path = _save(workbook, tmp_path / "broken-series.xlsx")

    scanned = _only_chart(ingestion.scan_workbook(str(path)))
    assert len(scanned["series"]) == 2
    valid, broken = scanned["series"]
    assert [point["value"] for point in valid["datapoints"]] == [100, 120, 150]
    assert broken["series_name"] == "Expenses"
    assert broken["value_range"] == broken_reference
    assert not broken["datapoints"] or all(
        point.get("value") is None for point in broken["datapoints"]
    )

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Financial workbook.",
            "tables": [
                {
                    "sheet_name": "Data",
                    "table_name": "Financial Data",
                    "cell_range": "A1:C4",
                    "column_headers": ["Month", "Revenue", "Expenses"],
                    "description": "Revenue and expenses by month.",
                }
            ],
            "visual_semantics": [],
            "key_findings": [],
        },
    )
    assert len(_chunks_of_type(chunks, "tabular_record")) == 3
    rendered_valid_points = "\n".join(
        chunk["text"] for chunk in _chunks_of_type(chunks, "chart_datapoint")
    )
    assert "Series: Revenue" in rendered_valid_points
    assert "Category: Mar" in rendered_valid_points
    assert "Value: 150" in rendered_valid_points


def test_embedded_image_is_discovered_and_gets_metadata_chunk(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    png = BytesIO()
    PillowImage.new("RGB", (3, 2), (255, 0, 0)).save(png, format="PNG")
    png.seek(0)
    sheet.add_image(ExcelImage(png), "H4")
    path = _save(workbook, tmp_path / "image.xlsx")

    scan = ingestion.scan_workbook(str(path))
    assert len(scan["images"]) == 1
    image = scan["images"][0]
    assert image["visual_id"] == "Dashboard:image:0"
    assert image["visual_type"] == "image"
    assert image["sheet_name"] == "Dashboard"
    assert image["location"] == "H4"

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Dashboard workbook.",
            "tables": [],
            "visual_semantics": [],
            "key_findings": [],
        },
    )
    visual_chunks = _chunks_of_type(chunks, "visual_metadata")
    assert len(visual_chunks) == 1
    assert visual_chunks[0]["content_type"] == "metadata"
    assert "Dashboard:image:0" in visual_chunks[0]["text"]
    assert "Sheet: Dashboard" in visual_chunks[0]["text"]
    assert "Location: H4" in visual_chunks[0]["text"]


def test_mocked_image_semantics_are_marked_vision_derived(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    png = BytesIO()
    PillowImage.new("RGB", (3, 2), (255, 0, 0)).save(png, format="PNG")
    png.seek(0)
    sheet.add_image(ExcelImage(png), "H4")
    path = _save(workbook, tmp_path / "vision-image.xlsx")

    def visual_analyzer(_path, visual, **_kwargs):
        return {
            "visual_id": visual["visual_id"],
            "visual_type": "dashboard",
            "name": "Operations Dashboard",
            "description": "A pasted operational dashboard.",
            "visible_text": ["Quarterly results"],
            "key_observations": ["The trend appears positive."],
            "datapoints": [{"label": "untrusted", "value": 999}],
            "confidence": 0.91,
        }

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Dashboard workbook.",
            "tables": [],
            "visual_semantics": [],
            "key_findings": [],
        },
        visual_analyzer=visual_analyzer,
    )
    visual = _chunks_of_type(chunks, "visual_metadata")[0]
    assert "Operations Dashboard" in visual["text"]
    assert "Visual Type: dashboard" in visual["text"]
    assert "Visible Text: Quarterly results" in visual["text"]
    assert "Observation: The trend appears positive." in visual["text"]
    assert "Vision Confidence: 0.91" in visual["text"]
    assert "Reliability: vision-derived semantics" in visual["text"]
    assert "999" not in visual["text"]


def test_vision_request_mime_matches_openpyxl_converted_image_bytes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    png = BytesIO()
    PillowImage.new("RGB", (3, 2), (12, 34, 56)).save(png, format="PNG")
    png.seek(0)
    sheet.add_image(ExcelImage(png), "H4")
    path = _save(workbook, tmp_path / "bitmap-image.xlsx")

    # Simulate an externally-authored XLSX that embeds BMP directly.  openpyxl
    # normally normalizes BMP to PNG at workbook-save time, so a package-level
    # replacement is needed to exercise the real interoperability path.
    bitmap = BytesIO()
    PillowImage.new("RGB", (3, 2), (12, 34, 56)).save(bitmap, format="BMP")
    with zipfile.ZipFile(path, "r") as package:
        content_types = package.read("[Content_Types].xml").replace(
            b"</Types>",
            b'<Default Extension="bmp" ContentType="image/bmp"/></Types>',
        )
        rel_name = next(
            name
            for name in package.namelist()
            if name.startswith("xl/drawings/_rels/")
        )
        relationships = package.read(rel_name).replace(b"image1.png", b"image1.bmp")
    _replace_zip_members(
        path,
        {
            "[Content_Types].xml": content_types,
            rel_name: relationships,
            "xl/media/image1.bmp": bitmap.getvalue(),
        },
        removed={"xl/media/image1.png"},
    )

    visual = ingestion.scan_workbook(str(path))["images"][0]
    assert visual["format"] == "bmp"

    monkeypatch.setattr(ingestion, "SPREADSHEET_VISION_MODEL", "vision-test")
    fake_client = FakeOpenAIClient(
        {
            "visual_id": "Dashboard:image:0",
            "visual_type": "screenshot",
            "name": "Bitmap",
            "description": "Converted embedded bitmap.",
            "visible_text": [],
            "key_observations": [],
            "confidence": 0.9,
        }
    )
    result = ingestion.analyze_embedded_visual_with_llm(
        str(path), visual, client=fake_client
    )

    assert result["visual_id"] == "Dashboard:image:0"
    image_url = fake_client.calls[0]["messages"][1]["content"][1]["image_url"][
        "url"
    ]
    # openpyxl converts unsupported-for-OOXML source encodings such as BMP to
    # PNG when `_data()` is read, so the MIME must describe those actual bytes.
    assert image_url.startswith("data:image/png;base64,iVBOR")


def test_image_analysis_failure_does_not_drop_image_or_chart(tmp_path: Path) -> None:
    path = _line_workbook(tmp_path / "visual-failure.xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook["Data"]
    png = BytesIO()
    PillowImage.new("RGB", (2, 2), (0, 0, 255)).save(png, format="PNG")
    png.seek(0)
    sheet.add_image(ExcelImage(png), "H4")
    workbook.save(path)

    def failing_analyzer(*_args, **_kwargs):
        raise RuntimeError("vision unavailable")

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis=_analysis_for_chart(),
        visual_analyzer=failing_analyzer,
    )
    assert len(_chunks_of_type(chunks, "visual_metadata")) == 1
    assert len(_chunks_of_type(chunks, "chart_metadata")) == 1
    assert len(_chunks_of_type(chunks, "chart_datapoint")) == 3


def test_unknown_llm_visual_id_cannot_replace_deterministic_chart(
    tmp_path: Path,
) -> None:
    path = _line_workbook(tmp_path / "semantics.xlsx")
    invalid_analysis = {
        "spreadsheet_summary": "A valid summary.",
        "tables": "not a list",
        "visual_semantics": [
            {
                "visual_id": "Invented:chart:99",
                "name": "Invented chart",
                "description": "This object is not in the workbook.",
                "x_axis_semantic": "Fabricated X",
                "y_axis_semantic": "Fabricated Y",
                "unit": "USD",
                "arbitrary_untrusted_key": "must not leak",
            }
        ],
        "key_findings": ["A finding.", "A finding.", 7],
        "arbitrary_top_level_key": "must not leak",
    }

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis=invalid_analysis,
    )

    chart_chunks = _chunks_of_type(chunks, "chart_metadata")
    assert len(chart_chunks) == 1
    assert "Monthly Revenue" in chart_chunks[0]["text"]
    rendered = "\n".join(chunk["text"] for chunk in chunks)
    assert "Invented chart" not in rendered
    assert "arbitrary_untrusted_key" not in rendered
    assert "arbitrary_top_level_key" not in rendered
    workbook_metadata = "\n".join(
        chunk["text"] for chunk in _chunks_of_type(chunks, "workbook_metadata")
    )
    assert workbook_metadata.count("A finding.") == 1


def test_llm_semantics_cannot_overwrite_deterministic_chart_facts_or_duplicate_chunks(
    tmp_path: Path,
) -> None:
    path = _line_workbook(tmp_path / "authoritative.xlsx")
    hostile_semantic = {
        "visual_id": "Data:chart:0",
        "name": "Revenue Semantics",
        "description": "A useful semantic description.",
        "x_axis_semantic": "Month",
        "y_axis_semantic": "Revenue",
        "unit": "USD",
        # These are outside the trusted semantic schema and must be ignored.
        "chart_type": "PieChart",
        "location": "Z99",
        "series": [{"series_name": "Invented"}],
        "datapoints": [{"category": "Mar", "value": 999_999}],
    }
    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        analysis={
            "spreadsheet_summary": "Revenue workbook.",
            "tables": [],
            "visual_semantics": [hostile_semantic, hostile_semantic],
            "key_findings": [],
        },
    )

    metadata = _chunks_of_type(chunks, "chart_metadata")
    datapoints = _chunks_of_type(chunks, "chart_datapoint")
    assert len(metadata) == 1
    assert len(datapoints) == 3
    rendered = "\n".join(chunk["text"] for chunk in metadata + datapoints)
    assert "Chart Type: LineChart" in rendered
    assert "Sheet: Data" in rendered
    assert "Value: 150" in rendered
    assert "PieChart" not in rendered
    assert "Z99" not in rendered
    assert "999999" not in rendered
    assert "Invented" not in rendered


def test_logically_identical_chart_points_are_deduplicated(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Month", "Revenue"])
    sheet.append(["Jan", 100])
    sheet.append(["Jan", 100])
    chart = LineChart()
    chart.title = "Monthly Revenue"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "E2")
    path = _save(workbook, tmp_path / "duplicate-points.xlsx")

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path), path.name, analysis=_analysis_for_chart()
    )

    datapoints = _chunks_of_type(chunks, "chart_datapoint")
    assert len(datapoints) == 1
    assert "Series: Revenue" in datapoints[0]["text"]
    assert "Category: Jan" in datapoints[0]["text"]
    assert "Value: 100" in datapoints[0]["text"]


def test_llm_input_limit_does_not_limit_deterministic_chart_scanning(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Large"
    sheet.append(["Category", "Value", "Notes"])
    for index in range(1, 151):
        sheet.append([f"row-{index}", index, f"note-{index}-" + ("z" * 500)])
    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=151),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=151))
    sheet.add_chart(chart, "D2")
    path = _save(workbook, tmp_path / "large.xlsx")
    monkeypatch.setattr(ingestion, "MAX_SPREADSHEET_LLM_CHARS", 500)
    fake_client = FakeOpenAIClient(
        {
            "spreadsheet_summary": "Large workbook.",
            "tables": [],
            "visual_semantics": [],
            "key_findings": [],
        }
    )

    scan = ingestion.scan_workbook(str(path))
    analysis = ingestion.analyze_spreadsheet_with_llm(
        str(path),
        scan,
        client=fake_client,
    )
    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path), path.name, analysis=analysis
    )

    user_prompt = fake_client.calls[0]["messages"][-1]["content"]
    assert len(user_prompt) < 20_000
    assert "note-150-" not in user_prompt
    assert len(_only_chart(scan)["series"][0]["datapoints"]) == 150
    assert len(_chunks_of_type(chunks, "chart_datapoint")) == 150


def test_ingestion_notes_reach_spreadsheet_llm_prompt(tmp_path: Path) -> None:
    path = _line_workbook(tmp_path / "noted.xlsx")
    fake_client = FakeOpenAIClient(_analysis_for_chart())
    notes = 'Yellow cells indicate "pending".\nRows 1-3 are KPI cards.'

    ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        client=fake_client,
        ingestion_notes=notes,
    )

    prompt = fake_client.calls[0]["messages"][-1]["content"]
    assert "BEGIN USER-PROVIDED INGESTION NOTES (JSON STRING)" in prompt
    assert json.dumps(notes) in prompt
    assert notes not in prompt
    assert "END USER-PROVIDED INGESTION NOTES" in prompt
    assert "deterministic workbook facts" in prompt
    system_prompt = fake_client.calls[0]["messages"][0]["content"]
    assert "untrusted contextual data" in system_prompt
    assert (
        "Deterministically extracted workbook facts are authoritative"
        in system_prompt
    )


def test_ingestion_notes_leave_room_for_workbook_structure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _line_workbook(tmp_path / "bounded-notes.xlsx")
    fake_client = FakeOpenAIClient(_analysis_for_chart())
    monkeypatch.setattr(ingestion, "MAX_SPREADSHEET_LLM_CHARS", 1_000)
    oversized_notes = "A" * 4_000

    ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        client=fake_client,
        ingestion_notes=oversized_notes,
    )

    prompt = fake_client.calls[0]["messages"][-1]["content"]
    assert "BEGIN USER-PROVIDED INGESTION NOTES (JSON STRING)" in prompt
    assert oversized_notes not in prompt
    assert "A1=Month" in prompt


def test_visual_llm_limit_does_not_drop_deterministic_chart_chunks(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _line_workbook(tmp_path / "visual-limit.xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook["Data"]
    second_chart = LineChart()
    second_chart.title = "Revenue Copy"
    second_chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    second_chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=4))
    sheet.add_chart(second_chart, "E20")
    workbook.save(path)

    monkeypatch.setattr(ingestion, "MAX_VISUALS_FOR_LLM", 1)
    fake_client = FakeOpenAIClient(
        {
            "spreadsheet_summary": "Two-chart workbook.",
            "tables": [],
            "visual_semantics": [],
            "key_findings": [],
        }
    )
    manifest = ingestion.scan_workbook(str(path))
    assert [chart["visual_id"] for chart in manifest["charts"]] == [
        "Data:chart:0",
        "Data:chart:1",
    ]
    analysis = ingestion.analyze_spreadsheet_with_llm(
        str(path), manifest, client=fake_client
    )
    prompt = fake_client.calls[0]["messages"][-1]["content"]
    assert "Data:chart:0" in prompt
    assert "Data:chart:1" not in prompt

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path), path.name, analysis=analysis
    )
    assert len(_chunks_of_type(chunks, "chart_metadata")) == 2
    assert len(_chunks_of_type(chunks, "chart_datapoint")) == 6


def test_llm_character_limit_also_bounds_large_visual_manifest(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    workbook.active.title = "Data"
    path = _save(workbook, tmp_path / "large-manifest.xlsx")
    monkeypatch.setattr(ingestion, "MAX_SPREADSHEET_LLM_CHARS", 500)
    fake_client = FakeOpenAIClient(
        {
            "spreadsheet_summary": "Workbook.",
            "tables": [],
            "visual_semantics": [],
            "key_findings": [],
        }
    )
    huge_manifest = {
        "sheets": [{"sheet_name": "Data", "max_row": 1, "max_column": 1}],
        "images": [],
        "charts": [
            {
                "visual_id": "Data:chart:0",
                "visual_type": "chart",
                "sheet_name": "Data",
                "chart_type": "LineChart",
                "series": [
                    {
                        "series_index": index,
                        "series_name": f"series-{index}-" + ("x" * 1_000),
                        "value_range": "'Data'!$A$1",
                        "datapoints": [],
                    }
                    for index in range(100)
                ],
            }
        ],
    }

    ingestion.analyze_spreadsheet_with_llm(
        str(path), huge_manifest, client=fake_client
    )

    prompt = fake_client.calls[0]["messages"][-1]["content"]
    assert len(prompt) < 20_000
    assert "series-99-" not in prompt


def test_malformed_llm_json_still_generates_deterministic_chart_chunks(
    tmp_path: Path,
) -> None:
    path = _line_workbook(tmp_path / "malformed-analysis.xlsx")

    chunks = ingestion.build_spreadsheet_chunk_specs(
        str(path),
        path.name,
        client=FakeRawOpenAIClient("not valid JSON"),
    )

    assert len(_chunks_of_type(chunks, "chart_metadata")) == 1
    datapoints = _chunks_of_type(chunks, "chart_datapoint")
    assert len(datapoints) == 3
    assert any(
        "Series: Revenue" in chunk["text"]
        and "Category: Mar" in chunk["text"]
        and "Value: 150" in chunk["text"]
        for chunk in datapoints
    )
    # The deterministic raw-row fallback remains useful even when semantic
    # table detection fails completely.
    assert len(_chunks_of_type(chunks, "tabular_record")) == 4
